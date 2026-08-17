#!/usr/bin/env python3
"""
Script 1 of 4 — Unpaywall Downloader
======================================
Reads DOIs from your Scopus Excel file.
Queries the Unpaywall API (free, no key needed — just an email).
Downloads the best open-access PDF found.
Names every file:  <S.No>_<sanitised title>.pdf

Usage
-----
    python 01_unpaywall_downloader.py

Environment overrides (optional):
    SCOPUS_EXCEL   path to your .xlsx file   (default: scopus_articale_only.xlsx)
    SCOPUS_SHEET   sheet name                 (default: researchArticalOnly)
    UNPAYWALL_EMAIL  registered email         (default: amanpalpalrayat@gmail.com)

Outputs
-------
    Research_Articles/PDFs/          downloaded PDFs
    Research_Articles/Reports/       01_unpaywall_report.xlsx  +  .csv
"""

import csv
import os
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote

import requests


# ── Configuration ─────────────────────────────────────────────────────────────

SCOPUS_EXCEL  = Path(os.getenv("SCOPUS_EXCEL",  "scopus_articale_only.xlsx"))
SCOPUS_SHEET  = os.getenv("SCOPUS_SHEET",  "Sheet1")
UNPAYWALL_EMAIL = os.getenv("UNPAYWALL_EMAIL", "amanpalpalrayat@gmail.com")

ROOT       = Path("Research_Articles")
PDF_DIR    = ROOT / "PDFs"
REPORT_DIR = ROOT / "Reports"

TIMEOUT    = 30
MAX_RETRIES = 3

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "UnpaywallDownloader/1.0 (academic research)",
    "Accept-Language": "en-US,en;q=0.9",
})

REPORT_FIELDS = [
    "paper_id", "doi", "title", "year", "source_title",
    "oa_status", "oa_host_type",
    "pdf_url_tried", "download_status", "local_pdf", "error",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def create_dirs():
    for d in (PDF_DIR, REPORT_DIR):
        d.mkdir(parents=True, exist_ok=True)


def clean_doi(v):
    v = str(v or "").strip().replace("\ufeff", "")
    v = re.sub(r"^\s*doi\s*:\s*", "", v, flags=re.I)
    v = re.sub(r"^https?://(dx\.)?doi\.org/", "", v, flags=re.I)
    return v.strip(" <>[](){}'\".,;")


def valid_doi(v):
    return bool(re.match(r"^10\.\d{4,9}/\S+$", v, flags=re.I))


def safe_filename(v, limit=180):
    v = str(v or "").strip()
    v = re.sub(r'[\\/:*?"<>|]', "_", v)
    v = re.sub(r"\s+", " ", v)
    return v[:limit].strip()


def is_valid_pdf(path: Path) -> bool:
    try:
        with open(path, "rb") as f:
            return f.read(5) == b"%PDF-"
    except Exception:
        return False


def quarantine_broken():
    """Move any .pdf files that are not real PDFs to PDFs/_broken/."""
    if not PDF_DIR.exists():
        return
    broken_dir = PDF_DIR / "_broken"
    moved = []
    for pdf in PDF_DIR.glob("*.pdf"):
        if not is_valid_pdf(pdf):
            broken_dir.mkdir(parents=True, exist_ok=True)
            dest = broken_dir / pdf.name
            try:
                pdf.replace(dest)
                moved.append(pdf.name)
            except Exception:
                pass
    if moved:
        print(f"\nMoved {len(moved)} broken file(s) to PDFs/_broken/ for retry:")
        for n in moved:
            print(f"  - {n}")
        print()


def get_json(url, params=None):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = SESSION.get(url, params=params, timeout=TIMEOUT)
            if r.status_code == 404:
                return None, "404 Not Found"
            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", 30))
                print(f"    Rate limited — waiting {wait}s …")
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json(), None
        except Exception as e:
            if attempt == MAX_RETRIES:
                return None, str(e)
            time.sleep(2 ** attempt)
    return None, "Max retries exceeded"


def download_pdf(url: str, dest: Path):
    """
    Returns (ok: bool, message: str).
    Streams the response, verifies %PDF- magic bytes, saves to dest.
    """
    try:
        r = SESSION.get(url, stream=True, timeout=60, allow_redirects=True)
        r.raise_for_status()

        content_type = r.headers.get("content-type", "").lower()
        first = r.raw.read(5)

        if first != b"%PDF-" and "application/pdf" not in content_type:
            r.close()
            return False, f"Not a PDF (Content-Type: {content_type})"

        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as fh:
            fh.write(first)
            for chunk in r.iter_content(65536):
                if chunk:
                    fh.write(chunk)
        r.close()

        if dest.stat().st_size < 1000:
            dest.unlink(missing_ok=True)
            return False, "File too small — likely an error page"

        return True, "Downloaded"

    except Exception as e:
        dest.unlink(missing_ok=True)
        return False, str(e)


# ── Scopus Excel reader ───────────────────────────────────────────────────────

def load_scopus():
    try:
        import pandas as pd
    except ImportError:
        print("ERROR: pip install pandas openpyxl")
        sys.exit(1)

    if not SCOPUS_EXCEL.exists():
        print(f"ERROR: Excel file not found: {SCOPUS_EXCEL.resolve()}")
        sys.exit(1)

    df = pd.read_excel(SCOPUS_EXCEL, sheet_name=SCOPUS_SHEET)
    df.columns = [str(c).strip() for c in df.columns]

    for col in ("DOI", "S.No"):
        if col not in df.columns:
            print(f"ERROR: Required column '{col}' missing. Found: {list(df.columns)}")
            sys.exit(1)

    records = []
    for _, row in df.iterrows():
        doi = clean_doi(row.get("DOI", ""))
        if not doi or not valid_doi(doi):
            continue

        s_no = str(row.get("S.No", "")).strip().rstrip(".0") or ""
        idx  = str(row.get("index", "")).strip().rstrip(".0") or ""
        paper_id = s_no or idx
        if not paper_id:
            continue

        records.append({
            "paper_id":    paper_id,
            "doi":         doi,
            "title":       str(row.get("Title",        "")).strip(),
            "authors":     str(row.get("Authors",      "")).strip(),
            "year":        str(row.get("Year",         "")).strip().rstrip(".0"),
            "source_title":str(row.get("Source title", "")).strip(),
        })

    return records


# ── Unpaywall API ─────────────────────────────────────────────────────────────

def query_unpaywall(doi: str):
    """Returns the Unpaywall JSON object or None."""
    url = f"https://api.unpaywall.org/v2/{quote(doi, safe='')}?email={UNPAYWALL_EMAIL}"
    data, err = get_json(url)
    return data, err


def best_pdf_url(data: dict) -> tuple[str, str, str]:
    """
    Returns (pdf_url, oa_status, host_type).
    Prefers repository/preprint over publisher HTML pages.
    """
    if not data:
        return "", "", ""

    oa_status  = data.get("oa_status", "")
    best       = data.get("best_oa_location") or {}
    locations  = data.get("oa_locations", [])

    # Collect all PDF URLs from all locations, ranked by preference
    candidates = []

    for loc in ([best] + locations):
        url = loc.get("url_for_pdf") or loc.get("url", "")
        if url and url not in [c[0] for c in candidates]:
            candidates.append((url, loc.get("host_type", "")))

    for url, host in candidates:
        if url:
            return url, oa_status, host

    return "", oa_status, ""


# ── Report writers ─────────────────────────────────────────────────────────────

def write_report(rows: list, stem: str):
    csv_path   = REPORT_DIR / f"{stem}.csv"
    xlsx_path  = REPORT_DIR / f"{stem}.xlsx"

    # CSV
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=REPORT_FIELDS, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in REPORT_FIELDS})

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Unpaywall Results"
        ws.append(REPORT_FIELDS)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        for row in rows:
            ws.append([row.get(f, "") for f in REPORT_FIELDS])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wide = {"title": 60, "doi": 38, "pdf_url_tried": 65, "local_pdf": 65, "error": 50}
        for i, f in enumerate(REPORT_FIELDS, 1):
            ws.column_dimensions[get_column_letter(i)].width = wide.get(f, 20)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Summary sheet
        ss = wb.create_sheet("Summary")
        total = len(rows)
        downloaded = sum(1 for r in rows if r.get("local_pdf"))
        failed = total - downloaded
        ss.append(["Metric", "Count"])
        ss.append(["Total DOIs", total])
        ss.append(["Downloaded", downloaded])
        ss.append(["Failed / No OA PDF", failed])

        wb.save(xlsx_path)
        print(f"\nReport saved: {xlsx_path}")

    except ImportError:
        print(f"\nReport saved (CSV only — pip install openpyxl for Excel): {csv_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    create_dirs()
    quarantine_broken()

    records = load_scopus()
    total   = len(records)

    print()
    print("=" * 68)
    print("  UNPAYWALL DOWNLOADER")
    print(f"  {total} DOIs loaded from {SCOPUS_EXCEL.name}")
    print("=" * 68)

    rows = []

    for i, rec in enumerate(records, 1):
        doi      = rec["doi"]
        paper_id = rec["paper_id"]
        title    = rec["title"]

        pdf_name = f"{safe_filename(paper_id)}_{safe_filename(title)}.pdf"
        out_path = PDF_DIR / pdf_name

        print(f"\n[{i}/{total}] {doi}")
        print(f"  Paper ID : {paper_id}")
        print(f"  PDF name : {pdf_name}")

        row = {
            **rec,
            "pdf_url_tried":  "",
            "oa_status":      "",
            "oa_host_type":   "",
            "download_status": "",
            "local_pdf":      "",
            "error":          "",
        }

        # Skip if already downloaded and valid
        if out_path.exists() and is_valid_pdf(out_path):
            print("  ✓ Already downloaded.")
            row["local_pdf"]       = str(out_path)
            row["download_status"] = "Already downloaded"
            rows.append(row)
            continue

        print("  Querying Unpaywall …")
        data, err = query_unpaywall(doi)

        if err and not data:
            print(f"  ✗ Unpaywall error: {err}")
            row["error"]          = err
            row["download_status"] = "Unpaywall API error"
            rows.append(row)
            continue

        pdf_url, oa_status, host_type = best_pdf_url(data)

        row["oa_status"]    = oa_status
        row["oa_host_type"] = host_type

        if not pdf_url:
            msg = f"No OA PDF found (oa_status={oa_status})"
            print(f"  ✗ {msg}")
            row["download_status"] = msg
            rows.append(row)
            continue

        row["pdf_url_tried"] = pdf_url
        print(f"  OA status  : {oa_status} ({host_type})")
        print(f"  PDF URL    : {pdf_url}")
        print("  Downloading …")

        ok, msg = download_pdf(pdf_url, out_path)

        if ok:
            print(f"  ✓ Saved → {out_path.name}")
            row["local_pdf"]       = str(out_path)
            row["download_status"] = "Downloaded"
        else:
            print(f"  ✗ {msg}")
            row["download_status"] = f"Download failed: {msg}"
            row["error"]           = msg

        rows.append(row)
        time.sleep(0.5)   # be polite to Unpaywall

    write_report(rows, "01_unpaywall_report")

    downloaded = sum(1 for r in rows if r.get("local_pdf"))
    print()
    print("=" * 68)
    print(f"  Done.  Downloaded {downloaded}/{total}")
    print("=" * 68)
    print()


if __name__ == "__main__":
    main()
