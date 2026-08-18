#!/usr/bin/env python3
"""
DOI Research Article Downloader - Version 2.1
==============================================

Retrieval order:
    1. Unpaywall
    2. Consensus
    3. Google Scholar

Existing PDFs from previous versions are preserved and never deleted.

Important:
- Uses legitimate/publicly available full-text locations.
- Does not bypass paywalls, logins, CAPTCHAs, or access controls.
- Google Scholar is a final fallback.
- Consensus API is optional. Without an API key, Consensus search URLs
  are recorded for manual follow-up.
"""

import csv
import os
import re
import sys
import time
import difflib
from pathlib import Path
from urllib.parse import quote

import requests


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = Path("DOI_NO.txt")

ROOT = Path("Research_Articles")
PDF_DIR = ROOT / "PDFs"
FAIL_DIR = ROOT / "Failed"
REPORT_DIR = ROOT / "Reports"

# Put your email here or set the UNPAYWALL_EMAIL environment variable.
UNPAYWALL_EMAIL = os.getenv(
    "UNPAYWALL_EMAIL",
    "amanpalrayat@gmail.com"
)

# Optional official Consensus API key.
CONSENSUS_API_KEY = os.getenv("CONSENSUS_API_KEY", "")

# Google Scholar is deliberately the LAST retrieval method.
ENABLE_GOOGLE_SCHOLAR = True
DOWNLOAD_SCHOLAR_PDFS = True

# Scholar requests are intentionally slow to reduce rate-limit problems.
SCHOLAR_DELAY = 8

TIMEOUT = 30
MAX_RETRIES = 3

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "ResearchArticleDownloader/2.1 "
        "(academic literature retrieval)"
    ),
    "Accept-Language": "en-US,en;q=0.9",
})


FIELDS = [
    "record_no",
    "doi",
    "title",
    "authors",
    "year",
    "journal",
    "publisher",
    "abstract",

    # Existing-paper protection
    "existing_pdf",
    "existing_pdf_path",

    # Unpaywall
    "unpaywall_status",
    "oa_status",
    "oa_host_type",
    "unpaywall_pdf_url",

    # Consensus
    "consensus_status",
    "consensus_title",
    "consensus_url",
    "consensus_pdf_url",
    "consensus_citation_count",
    "consensus_study_type",
    "consensus_takeaway",

    # Google Scholar
    "scholar_status",
    "scholar_result_title",
    "scholar_result_url",
    "scholar_pdf_url",

    # Final retrieval
    "local_pdf",
    "final_source",
    "download_status",
    "error",

    # Manual links
    "google_scholar_url",
    "consensus_search_url",
]


# ============================================================
# HELPERS
# ============================================================

def clean_doi(value):
    value = (value or "").strip().replace("\ufeff", "")

    value = re.sub(
        r"^\s*doi\s*:\s*",
        "",
        value,
        flags=re.I
    )

    value = re.sub(
        r"^https?://(dx\.)?doi\.org/",
        "",
        value,
        flags=re.I
    )

    return value.strip(" <>[](){}'\".,;")


def valid_doi(value):
    return bool(
        re.match(
            r"^10\.\d{4,9}/\S+$",
            value,
            flags=re.I
        )
    )


def normalize_text(value):
    value = (value or "").lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def safe_filename(value, limit=180):
    value = value or "untitled"
    value = re.sub(
        r'[<>:"/\\|?*\x00-\x1f]',
        "_",
        value
    )
    value = re.sub(r"\s+", " ", value).strip().rstrip(". ")

    return (value or "untitled")[:limit]


def scholar_url(doi, title):
    return (
        "https://scholar.google.com/scholar?q="
        + quote(title or doi)
    )


def consensus_search_url(doi):
    return (
        "https://consensus.app/search/?q="
        + quote(doi)
    )


def create_directories():
    for folder in (
        ROOT,
        PDF_DIR,
        FAIL_DIR,
        REPORT_DIR
    ):
        folder.mkdir(
            parents=True,
            exist_ok=True
        )


def get_json(url, params=None, headers=None):
    error = ""

    for attempt in range(MAX_RETRIES):
        try:
            response = SESSION.get(
                url,
                params=params,
                headers=headers,
                timeout=TIMEOUT
            )

            if response.status_code == 200:
                return response.json(), ""

            error = (
                f"HTTP {response.status_code}"
            )

        except Exception as exc:
            error = str(exc)

        if attempt < MAX_RETRIES - 1:
            time.sleep(2 * (attempt + 1))

    return {}, error


# ============================================================
# CROSSREF
# ============================================================

def get_crossref_metadata(doi):
    url = (
        "https://api.crossref.org/works/"
        + quote(doi, safe="")
    )

    data, error = get_json(url)

    if not data:
        return {}, error

    message = data.get("message", {})

    authors = []

    for author in message.get("author", []):
        name = " ".join(
            x for x in (
                author.get("given", ""),
                author.get("family", "")
            )
            if x
        )

        if name:
            authors.append(name)

    published = (
        message.get("published-print")
        or message.get("published-online")
        or message.get("issued")
        or {}
    )

    date_parts = published.get(
        "date-parts",
        [[]]
    )

    year = (
        date_parts[0][0]
        if date_parts and date_parts[0]
        else ""
    )

    abstract = message.get(
        "abstract",
        ""
    )

    abstract = re.sub(
        r"<[^>]+>",
        " ",
        abstract
    )

    abstract = re.sub(
        r"\s+",
        " ",
        abstract
    ).strip()

    return {
        "title": (
            message.get("title", [""])[0]
            if message.get("title")
            else ""
        ),
        "authors": "; ".join(authors),
        "year": year,
        "journal": (
            message.get(
                "container-title",
                [""]
            )[0]
            if message.get("container-title")
            else ""
        ),
        "publisher": message.get(
            "publisher",
            ""
        ),
        "abstract": abstract,
    }, ""


# ============================================================
# UNPAYWALL - STAGE 1
# ============================================================

def get_unpaywall(doi):
    if "@" not in UNPAYWALL_EMAIL:
        return {}, "UNPAYWALL_EMAIL is not configured"

    url = (
        "https://api.unpaywall.org/v2/"
        + quote(doi, safe="")
    )

    return get_json(
        url,
        params={"email": UNPAYWALL_EMAIL}
    )


def get_unpaywall_pdf_candidates(data):
    candidates = []

    best = data.get(
        "best_oa_location"
    ) or {}

    locations = data.get(
        "oa_locations",
        []
    ) or []

    ordered = [best]

    for location in locations:
        if location is not best:
            ordered.append(location)

    seen = set()

    for location in ordered:
        if not isinstance(location, dict):
            continue

        pdf_url = location.get(
            "url_for_pdf"
        )

        if (
            pdf_url
            and pdf_url not in seen
        ):
            candidates.append(pdf_url)
            seen.add(pdf_url)

    return candidates


# ============================================================
# CONSENSUS - STAGE 2
# ============================================================

def consensus_search(doi, title):
    """
    Optional official Consensus API integration.

    If no API key is configured, this stage does NOT scrape Consensus.
    The script simply records the Consensus DOI search URL for manual
    follow-up and continues to Google Scholar.
    """

    if not CONSENSUS_API_KEY:
        return {}, "Consensus API key not configured"

    headers = {
        "x-api-key": CONSENSUS_API_KEY,
        "Accept": "application/json",
        "User-Agent": SESSION.headers["User-Agent"],
    }

    queries = [doi]

    if title:
        queries.append(
            '"' + title + '"'
        )

    last_error = ""

    for query in queries:

        data, error = get_json(
            "https://api.consensus.app/v1/quick_search",
            params={
                "query": query,
                "page": 0,
            },
            headers=headers
        )

        if data:
            return data, ""

        last_error = error

        time.sleep(1)

    return {}, last_error


def extract_consensus_results(data):
    if not data:
        return []

    results = (
        data.get("papers")
        or data.get("results")
        or data.get("data")
        or []
    )

    if isinstance(results, dict):
        results = [results]

    return [
        result
        for result in results
        if isinstance(result, dict)
    ]


def consensus_pdf_candidates(result):
    """
    Try only explicit/public PDF-like URLs returned by the API.
    We do not attempt to bypass publisher restrictions.
    """

    candidates = []

    for key in (
        "pdf_url",
        "full_text_url",
        "open_access_url",
        "url"
    ):
        value = result.get(key)

        if (
            isinstance(value, str)
            and value.startswith(("http://", "https://"))
            and value not in candidates
        ):
            candidates.append(value)

    return candidates


# ============================================================
# GOOGLE SCHOLAR - STAGE 3
# ============================================================

def scholar_search(doi, title):
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return {
            "status": "Playwright not installed",
            "found": False,
        }

    result = {
        "status": "No Scholar result",
        "found": False,
        "result_title": "",
        "result_url": "",
        "pdf_url": "",
    }

    with sync_playwright() as playwright:

        browser = playwright.chromium.launch(
            headless=False
        )

        page = browser.new_page(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/138.0.0.0 Safari/537.36"
            )
        )

        try:

            page.goto(
                scholar_url(doi, title),
                wait_until="domcontentloaded",
                timeout=60000
            )

            time.sleep(2)

            body = page.locator(
                "body"
            ).inner_text().lower()

            challenge_terms = [
                "captcha",
                "recaptcha",
                "unusual traffic",
                "not a robot",
                "our systems have detected",
            ]

            if any(
                term in body
                for term in challenge_terms
            ):
                print()
                print(
                    "Google Scholar verification "
                    "appeared."
                )
                print(
                    "Solve it manually in the "
                    "browser window."
                )

                input(
                    "Press ENTER after Scholar "
                    "is accessible..."
                )

            results = page.locator(
                "div.gs_ri"
            )

            count = results.count()

            for i in range(
                min(count, 5)
            ):

                item = results.nth(i)

                try:
                    result_title = (
                        item.locator(
                            "h3.gs_rt"
                        ).inner_text()
                    )
                except Exception:
                    result_title = ""

                try:
                    result_url = (
                        item.locator(
                            "h3.gs_rt a"
                        ).get_attribute("href")
                    )
                except Exception:
                    result_url = ""

                pdf_url = ""

                links = item.locator("a")

                for j in range(
                    links.count()
                ):
                    try:
                        text = (
                            links.nth(j)
                            .inner_text()
                            .strip()
                        )

                        href = (
                            links.nth(j)
                            .get_attribute("href")
                        )

                    except Exception:
                        continue

                    if not href:
                        continue

                    if (
                        "pdf" in text.lower()
                        or ".pdf" in href.lower()
                    ):
                        pdf_url = href
                        break

                score = difflib.SequenceMatcher(
                    None,
                    normalize_text(title)[:300],
                    normalize_text(result_title)[:300]
                ).ratio()

                if (
                    i == 0
                    or score >= 0.80
                ):

                    result.update({
                        "status":
                            "Scholar result found",
                        "found": True,
                        "result_title":
                            result_title,
                        "result_url":
                            result_url or "",
                        "pdf_url":
                            pdf_url,
                    })

                    break

        except Exception as exc:

            result["status"] = (
                f"Scholar error: {exc}"
            )

        finally:
            browser.close()

    return result


# ============================================================
# EXISTING PDF PROTECTION
# ============================================================

def load_previous_rows():
    candidates = [
        REPORT_DIR /
        "master_paper_database_progress.csv",

        REPORT_DIR /
        "download_report_progress.csv",

        REPORT_DIR /
        "download_report.csv",

        REPORT_DIR /
        "master_paper_database.csv",
    ]

    for report in candidates:

        if not report.exists():
            continue

        try:

            with open(
                report,
                encoding="utf-8-sig",
                newline=""
            ) as file:

                rows = csv.DictReader(file)

                result = {}

                for row in rows:

                    doi = clean_doi(
                        row.get("doi", "")
                    )

                    if doi:
                        result[
                            doi.lower()
                        ] = row

                if result:
                    print(
                        "Previous progress loaded:",
                        report
                    )

                return result

        except Exception:
            continue

    return {}


def find_existing_pdf(
    doi,
    title,
    previous
):

    old = previous.get(
        doi.lower(),
        {}
    )

    old_path = old.get(
        "local_pdf",
        ""
    )

    if (
        old_path
        and Path(old_path).exists()
    ):
        return (
            Path(old_path),
            "previous report"
        )

    if not title:
        return None, ""

    target = normalize_text(title)

    best = None
    best_score = 0

    if PDF_DIR.exists():

        for pdf in PDF_DIR.glob(
            "*.pdf"
        ):

            score = difflib.SequenceMatcher(
                None,
                target[:300],
                normalize_text(
                    pdf.stem
                )[:300]
            ).ratio()

            if score > best_score:
                best = pdf
                best_score = score

    if (
        best
        and best_score >= 0.92
    ):
        return (
            best,
            f"title match {best_score:.2f}"
        )

    return None, ""


# ============================================================
# PDF DOWNLOAD
# ============================================================

def download_pdf(
    url,
    output_path
):

    try:

        with SESSION.get(
            url,
            timeout=TIMEOUT,
            stream=True,
            allow_redirects=True
        ) as response:

            if response.status_code != 200:
                return (
                    False,
                    f"HTTP {response.status_code}",
                    response.url
                )

            content_type = (
                response.headers.get(
                    "Content-Type",
                    ""
                ).lower()
            )

            first = response.raw.read(5)

            if (
                "application/pdf"
                not in content_type
                and first != b"%PDF-"
            ):
                return (
                    False,
                    "Response is not a PDF",
                    response.url
                )

            output_path.parent.mkdir(
                parents=True,
                exist_ok=True
            )

            with open(
                output_path,
                "wb"
            ) as file:

                file.write(first)

                for chunk in response.iter_content(
                    65536
                ):

                    if chunk:
                        file.write(chunk)

        if output_path.stat().st_size < 1000:

            output_path.unlink(
                missing_ok=True
            )

            return (
                False,
                "Suspiciously small PDF",
                url
            )

        return (
            True,
            "Downloaded",
            response.url
        )

    except Exception as exc:

        return (
            False,
            str(exc),
            url
        )


# ============================================================
# REPORTING
# ============================================================

def write_csv(
    rows,
    path
):

    with open(
        path,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as file:

        writer = csv.DictWriter(
            file,
            fieldnames=FIELDS,
            extrasaction="ignore"
        )

        writer.writeheader()

        for row in rows:
            writer.writerow({
                field:
                    row.get(field, "")
                for field in FIELDS
            })


def write_excel(
    rows,
    path
):

    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font,
            PatternFill,
            Alignment
        )
        from openpyxl.utils import (
            get_column_letter
        )

    except ImportError:
        return False

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Master Database"

    sheet.append(FIELDS)

    for cell in sheet[1]:
        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )

    for row in rows:

        sheet.append([
            row.get(field, "")
            for field in FIELDS
        ])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        sheet.dimensions
    )

    wide = {
        "title": 55,
        "authors": 40,
        "abstract": 70,
        "unpaywall_pdf_url": 70,
        "consensus_url": 70,
        "consensus_pdf_url": 70,
        "scholar_result_url": 70,
        "scholar_pdf_url": 70,
        "local_pdf": 55,
        "error": 50,
    }

    for index, field in enumerate(
        FIELDS,
        start=1
    ):

        letter = get_column_letter(
            index
        )

        sheet.column_dimensions[
            letter
        ].width = wide.get(
            field,
            20
        )

    for row in sheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    summary = workbook.create_sheet(
        "Summary"
    )

    total = len(rows)

    preserved = sum(
        r.get("existing_pdf") == "Yes"
        for r in rows
    )

    downloaded = sum(
        r.get("download_status", "")
        .startswith("Downloaded")
        for r in rows
    )

    unpaywall = sum(
        r.get("final_source") == "Unpaywall"
        for r in rows
    )

    consensus = sum(
        r.get("final_source") == "Consensus"
        for r in rows
    )

    scholar = sum(
        r.get("final_source") == "Google Scholar"
        for r in rows
    )

    manual = sum(
        "manual"
        in r.get(
            "download_status",
            ""
        ).lower()
        or "not found"
        in r.get(
            "download_status",
            ""
        ).lower()
        for r in rows
    )

    summary_rows = [
        ("Total DOI records", total),
        ("Existing PDFs preserved", preserved),
        ("New PDFs downloaded", downloaded),
        ("Downloaded via Unpaywall", unpaywall),
        ("Downloaded via Consensus", consensus),
        ("Downloaded via Google Scholar", scholar),
        ("Manual/unresolved", manual),
    ]

    for item in summary_rows:
        summary.append(item)

    workbook.save(path)

    return True


# ============================================================
# MAIN
# ============================================================

def main():

    create_directories()

    if not INPUT_FILE.exists():

        print(
            "ERROR: DOI_NO(2).txt was not found."
        )

        print(
            "Place it beside this script."
        )

        sys.exit(1)

    # --------------------------------------------------------
    # DOI list
    # --------------------------------------------------------

    lines = INPUT_FILE.read_text(
        encoding="utf-8-sig",
        errors="ignore"
    ).splitlines()

    seen = set()
    dois = []
    invalid = 0

    for line in lines:

        doi = clean_doi(line)

        if not doi:
            continue

        if not valid_doi(doi):

            invalid += 1
            continue

        if doi.lower() not in seen:

            seen.add(
                doi.lower()
            )

            dois.append(doi)

    print()
    print("=" * 70)
    print(
        "DOI RESEARCH ARTICLE DOWNLOADER V2.1"
    )
    print("=" * 70)

    print(
        "Retrieval order:"
    )

    print(
        "  1. Unpaywall"
    )

    print(
        "  2. Consensus"
    )

    print(
        "  3. Google Scholar"
    )

    print()
    print(
        f"Unique valid DOIs: {len(dois)}"
    )

    previous = load_previous_rows()

    print(
        f"Previous records: {len(previous)}"
    )

    # --------------------------------------------------------
    # Scholar availability
    # --------------------------------------------------------

    scholar_enabled = (
        ENABLE_GOOGLE_SCHOLAR
    )

    if scholar_enabled:

        try:
            import playwright  # noqa
        except ImportError:

            print(
                "Playwright is not installed."
            )

            print(
                "Google Scholar stage will be skipped."
            )

            scholar_enabled = False

    rows = []

    # ========================================================
    # PROCESS DOI
    # ========================================================

    for number, doi in enumerate(
        dois,
        start=1
    ):

        print()
        print(
            f"[{number}/{len(dois)}] {doi}"
        )

        metadata, crossref_error = (
            get_crossref_metadata(doi)
        )

        time.sleep(0.5)

        old = previous.get(
            doi.lower(),
            {}
        )

        title = (
            metadata.get("title")
            or old.get("title", "")
        )

        row = {
            field: ""
            for field in FIELDS
        }

        row.update({
            "record_no": number,
            "doi": doi,
            "title": title,

            "authors":
                metadata.get(
                    "authors",
                    old.get("authors", "")
                ),

            "year":
                metadata.get(
                    "year",
                    old.get("year", "")
                ),

            "journal":
                metadata.get(
                    "journal",
                    old.get("journal", "")
                ),

            "publisher":
                metadata.get(
                    "publisher",
                    old.get("publisher", "")
                ),

            "abstract":
                metadata.get(
                    "abstract",
                    old.get("abstract", "")
                ),

            "google_scholar_url":
                scholar_url(
                    doi,
                    title
                ),

            "consensus_search_url":
                consensus_search_url(
                    doi
                ),

            "error":
                crossref_error,
        })

        # ====================================================
        # EXISTING PDF CHECK
        # ====================================================

        existing, existing_reason = (
            find_existing_pdf(
                doi,
                title,
                previous
            )
        )

        if existing:

            row.update({
                "existing_pdf":
                    "Yes",

                "existing_pdf_path":
                    str(existing),

                "local_pdf":
                    str(existing),

                "final_source":
                    "Existing",

                "download_status":
                    "Existing - preserved",
            })

            print(
                "  ✓ Existing PDF preserved:",
                existing
            )

            rows.append(row)

            write_csv(
                rows,
                REPORT_DIR /
                "master_paper_database_progress.csv"
            )

            continue

        # ====================================================
        # STAGE 1 - UNPAYWALL
        # ====================================================

        print(
            "  [1/3] Checking Unpaywall..."
        )

        upw, upw_error = (
            get_unpaywall(doi)
        )

        time.sleep(0.5)

        if upw:

            row["unpaywall_status"] = (
                "Record found"
            )

            row["oa_status"] = (
                upw.get(
                    "oa_status",
                    ""
                )
            )

            best = (
                upw.get(
                    "best_oa_location"
                )
                or {}
            )

            row["oa_host_type"] = (
                best.get(
                    "host_type",
                    ""
                )
            )

            candidates = (
                get_unpaywall_pdf_candidates(
                    upw
                )
            )

            downloaded = False

            for pdf_url in candidates:

                output = (
                    PDF_DIR
                    /
                    (
                        safe_filename(
                            f"{number:03d}_{title or doi}"
                        )
                        + ".pdf"
                    )
                )

                ok, message, final_url = (
                    download_pdf(
                        pdf_url,
                        output
                    )
                )

                if ok:

                    row.update({
                        "unpaywall_pdf_url":
                            pdf_url,

                        "local_pdf":
                            str(output),

                        "final_source":
                            "Unpaywall",

                        "download_status":
                            "Downloaded - Unpaywall",
                    })

                    print(
                        "  ✓ PDF downloaded from Unpaywall"
                    )

                    downloaded = True
                    break

                row["error"] = message

            if downloaded:

                rows.append(row)

                write_csv(
                    rows,
                    REPORT_DIR /
                    "master_paper_database_progress.csv"
                )

                continue

        else:

            row["unpaywall_status"] = (
                "No record"
            )

            if upw_error:
                row["error"] = (
                    "Unpaywall: "
                    + upw_error
                )

        # ====================================================
        # STAGE 2 - CONSENSUS
        # ====================================================

        print(
            "  [2/3] Checking Consensus..."
        )

        if CONSENSUS_API_KEY:

            consensus_data, consensus_error = (
                consensus_search(
                    doi,
                    title
                )
            )

            results = (
                extract_consensus_results(
                    consensus_data
                )
            )

            if results:

                # Find closest title match.
                best_result = results[0]

                best_score = 0

                for candidate in results[:10]:

                    candidate_title = (
                        candidate.get(
                            "title",
                            ""
                        )
                    )

                    score = (
                        difflib.SequenceMatcher(
                            None,
                            normalize_text(
                                title
                            )[:300],
                            normalize_text(
                                candidate_title
                            )[:300]
                        ).ratio()
                    )

                    if score > best_score:

                        best_score = score
                        best_result = candidate

                row.update({
                    "consensus_status":
                        "Result found",

                    "consensus_title":
                        best_result.get(
                            "title",
                            ""
                        ),

                    "consensus_url":
                        best_result.get(
                            "url",
                            ""
                        ),

                    "consensus_citation_count":
                        best_result.get(
                            "citation_count",
                            ""
                        ),

                    "consensus_study_type":
                        best_result.get(
                            "study_type",
                            ""
                        ),

                    "consensus_takeaway":
                        best_result.get(
                            "takeaway",
                            ""
                        ),
                })

                # Try explicit public/full-text URLs.
                for pdf_url in (
                    consensus_pdf_candidates(
                        best_result
                    )
                ):

                    output = (
                        PDF_DIR
                        /
                        (
                            safe_filename(
                                f"{number:03d}_{title or doi}"
                            )
                            + ".pdf"
                        )
                    )

                    ok, message, final_url = (
                        download_pdf(
                            pdf_url,
                            output
                        )
                    )

                    if ok:

                        row.update({
                            "consensus_pdf_url":
                                pdf_url,

                            "local_pdf":
                                str(output),

                            "final_source":
                                "Consensus",

                            "download_status":
                                "Downloaded - Consensus",
                        })

                        print(
                            "  ✓ PDF downloaded via Consensus"
                        )

                        break

                    row["error"] = message

            else:

                row["consensus_status"] = (
                    "No API match"
                )

                if consensus_error:

                    row["error"] = (
                        f"{row['error']} | "
                        f"Consensus: "
                        f"{consensus_error}"
                    ).strip(" |")

        else:

            row["consensus_status"] = (
                "Link only - API key not configured"
            )

            print(
                "  Consensus API key not configured."
            )

            print(
                "  Consensus search URL saved in report."
            )

        # ----------------------------------------------------
        # If Consensus supplied a PDF, stop here.
        # ----------------------------------------------------

        if row["local_pdf"]:

            rows.append(row)

            write_csv(
                rows,
                REPORT_DIR /
                "master_paper_database_progress.csv"
            )

            continue

        # ====================================================
        # STAGE 3 - GOOGLE SCHOLAR
        # ====================================================

        if scholar_enabled:

            print(
                "  [3/3] Checking Google Scholar..."
            )

            scholar_result = (
                scholar_search(
                    doi,
                    title
                )
            )

            row.update({
                "scholar_status":
                    scholar_result.get(
                        "status",
                        ""
                    ),

                "scholar_result_title":
                    scholar_result.get(
                        "result_title",
                        ""
                    ),

                "scholar_result_url":
                    scholar_result.get(
                        "result_url",
                        ""
                    ),

                "scholar_pdf_url":
                    scholar_result.get(
                        "pdf_url",
                        ""
                    ),
            })

            if (
                scholar_result.get(
                    "found"
                )
            ):

                if (
                    DOWNLOAD_SCHOLAR_PDFS
                    and scholar_result.get(
                        "pdf_url"
                    )
                ):

                    output = (
                        PDF_DIR
                        /
                        (
                            safe_filename(
                                f"{number:03d}_{title or doi}"
                            )
                            + ".pdf"
                        )
                    )

                    ok, message, final_url = (
                        download_pdf(
                            scholar_result[
                                "pdf_url"
                            ],
                            output
                        )
                    )

                    if ok:

                        row.update({
                            "local_pdf":
                                str(output),

                            "final_source":
                                "Google Scholar",

                            "download_status":
                                "Downloaded - Google Scholar",
                        })

                        print(
                            "  ✓ PDF downloaded from Google Scholar"
                        )

                    else:

                        row["error"] = (
                            f"{row['error']} | "
                            f"Scholar PDF: "
                            f"{message}"
                        ).strip(" |")

                if not row["local_pdf"]:

                    row["download_status"] = (
                        "Scholar result - manual review"
                    )

            else:

                row["download_status"] = (
                    "Full text not found"
                )

            time.sleep(
                SCHOLAR_DELAY
            )

        else:

            row["scholar_status"] = (
                "Disabled or Playwright unavailable"
            )

            row["download_status"] = (
                "Full text not found"
            )

        # ====================================================
        # SAVE AFTER EVERY PAPER
        # ====================================================

        rows.append(row)

        write_csv(
            rows,
            REPORT_DIR /
            "master_paper_database_progress.csv"
        )

    # ========================================================
    # FINAL REPORTS
    # ========================================================

    master_csv = (
        REPORT_DIR /
        "master_paper_database.csv"
    )

    master_xlsx = (
        REPORT_DIR /
        "master_paper_database.xlsx"
    )

    write_csv(
        rows,
        master_csv
    )

    excel_created = write_excel(
        rows,
        master_xlsx
    )

    total = len(rows)

    preserved = sum(
        r["existing_pdf"] == "Yes"
        for r in rows
    )

    downloaded = sum(
        r["download_status"].startswith(
            "Downloaded"
        )
        for r in rows
    )

    from_unpaywall = sum(
        r["final_source"] == "Unpaywall"
        for r in rows
    )

    from_consensus = sum(
        r["final_source"] == "Consensus"
        for r in rows
    )

    from_scholar = sum(
        r["final_source"]
        == "Google Scholar"
        for r in rows
    )

    manual = sum(
        (
            "manual"
            in r["download_status"].lower()
        )
        or
        (
            "not found"
            in r["download_status"].lower()
        )
        for r in rows
    )

    summary = f"""
DOI RESEARCH ARTICLE DOWNLOADER V2.1
====================================

Retrieval order:
1. Unpaywall
2. Consensus
3. Google Scholar

Total unique DOI records:
{total}

Existing PDFs preserved:
{preserved}

New PDFs downloaded:
{downloaded}

Downloaded via Unpaywall:
{from_unpaywall}

Downloaded via Consensus:
{from_consensus}

Downloaded via Google Scholar:
{from_scholar}

Manual/unresolved:
{manual}

Excel report created:
{excel_created}

PDF folder:
{PDF_DIR.resolve()}

Reports:
{REPORT_DIR.resolve()}

Important:
Existing PDFs are never deleted or deliberately re-downloaded.
Google Scholar CAPTCHA/access restrictions are not bypassed.
Publisher paywalls are not bypassed.
"""

    (
        REPORT_DIR /
        "SUMMARY.txt"
    ).write_text(
        summary.strip()
        + "\n",
        encoding="utf-8"
    )

    print()
    print("=" * 70)
    print(
        "V2.1 COMPLETED"
    )
    print("=" * 70)
    print(
        f"Existing PDFs preserved : {preserved}"
    )
    print(
        f"New PDFs downloaded     : {downloaded}"
    )
    print(
        f"  Unpaywall             : {from_unpaywall}"
    )
    print(
        f"  Consensus             : {from_consensus}"
    )
    print(
        f"  Google Scholar        : {from_scholar}"
    )
    print(
        f"Manual/unresolved       : {manual}"
    )
    print()
    print(
        "Reports:",
        REPORT_DIR.resolve()
    )


if __name__ == "__main__":
    main()
