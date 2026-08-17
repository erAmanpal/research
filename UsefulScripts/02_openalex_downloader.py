#!/usr/bin/env python3
"""
Script 4 of 4 — OpenAlex (Extended Legal OA) Downloader
=========================================================
Reads DOIs from your Scopus Excel file.
Queries the OpenAlex API (free, no key needed) by DOI.
Tries every open-access PDF URL OpenAlex knows about for that DOI.
Names every file:  <S.No>_<sanitised title>.pdf

OpenAlex indexes > 250 million works and aggregates OA locations from
Unpaywall, institutional repositories, preprint servers (arXiv, bioRxiv,
SSRN), and publisher gold-OA outlets — often finding PDFs that Unpaywall
alone misses because OpenAlex re-checks more frequently.

Usage
-----
    python 04_openalex_downloader.py

Environment overrides (optional):
    SCOPUS_EXCEL    path to your .xlsx file   (default: scopus_articale_only.xlsx)
    SCOPUS_SHEET    sheet name                 (default: researchArticalOnly)
    OPENALEX_EMAIL  polite-pool email          (default: amanpalpalrayat@gmail.com)
                    (speeds up API responses — see https://docs.openalex.org)

Outputs
-------
    Research_Articles/PDFs/          downloaded PDFs
    Research_Articles/Reports/       04_openalex_report.xlsx  +  .csv
"""

import csv
import os
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote

import requests


# ── Configuration ─────────────────────────────────────────────────────────────

SCOPUS_EXCEL   = Path(os.getenv("SCOPUS_EXCEL", "scopus_articale_only.xlsx"))
SCOPUS_SHEET   = os.getenv("SCOPUS_SHEET", "Sheet1")
OPENALEX_EMAIL = os.getenv("OPENALEX_EMAIL", "amanpalpalrayat@gmail.com")

ROOT       = Path("Research_Articles")
PDF_DIR    = ROOT / "PDFs"
REPORT_DIR = ROOT / "Reports"

TIMEOUT    = 30
MAX_RETRIES = 3

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "OpenAlexDownloader/1.0 (academic research)",
    "Accept-Language": "en-US,en;q=0.9",
})

REPORT_FIELDS = [
    "paper_id", "doi", "title", "year", "source_title",
    "oa_status", "oa_url_count",
    "pdf_url_tried", "all_oa_urls",
    "download_status", "local_pdf", "error",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def create_dirs():
    for d in (PDF_DIR, REPORT_DIR):
        d.mkdir(parents=True, exist_ok=True)


def clean_doi(v):
    v = str(v or "").strip().replace("\ufeff", "")
    v = re.sub(r"^\s*doi\s*:\s*", "", v, flags=re.I)
    v = re.sub(r"^https?://(dx\.)?doi\.org/", "", v, flags=re.I)
    return v.strip(" <>[](){}'\".,;")


def valid_doi(v):
    return bool(re.match(r"^10\.\d{4,9}/\S+$", v, flags=re.I))


def safe_filename(v, limit=180):
    v = str(v or "").strip()
    v = re.sub(r'[\\/:*?"<>|]', "_", v)
    v = re.sub(r"\s+", " ", v)
    return v[:limit].strip()


def is_valid_pdf(path: Path) -> bool:
    try:
        with open(path, "rb") as f:
            return f.read(5) == b"%PDF-"
    except Exception:
        return False


def quarantine_broken():
    if not PDF_DIR.exists():
        return
    broken_dir = PDF_DIR / "_broken"
    moved = []
    for pdf in PDF_DIR.glob("*.pdf"):
        if not is_valid_pdf(pdf):
            broken_dir.mkdir(parents=True, exist_ok=True)
            dest = broken_dir / pdf.name
            try:
                pdf.replace(dest)
                moved.append(pdf.name)
            except Exception:
                pass
    if moved:
        print(f"\nMoved {len(moved)} broken file(s) to PDFs/_broken/ for retry:")
        for n in moved:
            print(f"  - {n}")
        print()


def get_json(url, params=None):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = SESSION.get(url, params=params, timeout=TIMEOUT)
            if r.status_code == 404:
                return None, "404 Not Found"
            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", 30))
                print(f"    Rate limited — waiting {wait}s …")
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json(), None
        except Exception as e:
            if attempt == MAX_RETRIES:
                return None, str(e)
            time.sleep(2 ** attempt)
    return None, "Max retries exceeded"


def download_pdf(url: str, dest: Path) -> tuple[bool, str]:
    """Download and verify %PDF- magic bytes. Returns (ok, message)."""
    try:
        r = SESSION.get(url, stream=True, timeout=60, allow_redirects=True)
        r.raise_for_status()

        ct    = r.headers.get("content-type", "").lower()
        first = r.raw.read(5)

        if first != b"%PDF-" and "application/pdf" not in ct:
            r.close()
            return False, f"Not a PDF (Content-Type: {ct})"

        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as fh:
            fh.write(first)
            for chunk in r.iter_content(65536):
                if chunk:
                    fh.write(chunk)
        r.close()

        if dest.stat().st_size < 1000:
            dest.unlink(missing_ok=True)
            return False, "File too small"

        return True, "Downloaded"

    except Exception as e:
        dest.unlink(missing_ok=True)
        return False, str(e)


# ── Scopus Excel reader ───────────────────────────────────────────────────────

def load_scopus():
    try:
        import pandas as pd
    except ImportError:
        print("ERROR: pip install pandas openpyxl")
        sys.exit(1)

    if not SCOPUS_EXCEL.exists():
        print(f"ERROR: Excel file not found: {SCOPUS_EXCEL.resolve()}")
        sys.exit(1)

    df = pd.read_excel(SCOPUS_EXCEL, sheet_name=SCOPUS_SHEET)
    df.columns = [str(c).strip() for c in df.columns]

    for col in ("DOI", "S.No"):
        if col not in df.columns:
            print(f"ERROR: Column '{col}' missing. Found: {list(df.columns)}")
            sys.exit(1)

    records = []
    for _, row in df.iterrows():
        doi = clean_doi(row.get("DOI", ""))
        if not doi or not valid_doi(doi):
            continue
        s_no = str(row.get("S.No",  "")).strip().rstrip(".0") or ""
        idx  = str(row.get("index", "")).strip().rstrip(".0") or ""
        paper_id = s_no or idx
        if not paper_id:
            continue
        records.append({
            "paper_id":     paper_id,
            "doi":          doi,
            "title":        str(row.get("Title",        "")).strip(),
            "authors":      str(row.get("Authors",      "")).strip(),
            "year":         str(row.get("Year",         "")).strip().rstrip(".0"),
            "source_title": str(row.get("Source title", "")).strip(),
        })
    return records


# ── OpenAlex API ──────────────────────────────────────────────────────────────

def query_openalex(doi: str) -> tuple[dict | None, str | None]:
    """
    Fetch the OpenAlex work record for a DOI.
    Uses the polite pool (mailto= parameter) for faster responses.
    """
    url    = f"https://api.openalex.org/works/https://doi.org/{quote(doi, safe='')}"
    params = {"mailto": OPENALEX_EMAIL}
    return get_json(url, params=params)


def extract_pdf_candidates(data: dict) -> list[dict]:
    """
    Return a ranked list of OA PDF candidate dicts:
        { url, host_type, version, is_best }
    Ordered: best_oa_location first, then all other oa_locations,
    preferring published > acceptedVersion > submittedVersion.
    """
    if not data:
        return []

    oa = data.get("open_access", {})
    if not oa.get("is_oa"):
        return []

    best_loc    = data.get("best_oa_location") or {}
    all_locs    = data.get("locations", [])

    version_rank = {"publishedVersion": 0, "acceptedVersion": 1,
                    "submittedVersion": 2, "": 9}

    candidates = []

    def add(loc, is_best=False):
        pdf_url = loc.get("pdf_url") or ""
        if not pdf_url:
            return
        candidates.append({
            "url":       pdf_url,
            "host_type": loc.get("source", {}).get("type", "") or loc.get("host_type", ""),
            "version":   loc.get("version", ""),
            "is_best":   is_best,
            "_rank":     version_rank.get(loc.get("version", ""), 9),
        })

    add(best_loc, is_best=True)

    for loc in all_locs:
        pdf_url = loc.get("pdf_url") or ""
        if pdf_url and pdf_url != best_loc.get("pdf_url"):
            add(loc)

    # Stable sort: best_oa first, then by version rank
    candidates.sort(key=lambda c: (0 if c["is_best"] else 1, c["_rank"]))

    # Deduplicate
    seen  = set()
    dedup = []
    for c in candidates:
        if c["url"] not in seen:
            seen.add(c["url"])
            dedup.append(c)

    return dedup


# ── Report writers ─────────────────────────────────────────────────────────────

def write_report(rows, stem):
    csv_path  = REPORT_DIR / f"{stem}.csv"
    xlsx_path = REPORT_DIR / f"{stem}.xlsx"

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=REPORT_FIELDS, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in REPORT_FIELDS})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "OpenAlex Results"
        ws.append(REPORT_FIELDS)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        for row in rows:
            ws.append([row.get(f, "") for f in REPORT_FIELDS])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wide = {
            "title": 60, "doi": 38,
            "pdf_url_tried": 65, "all_oa_urls": 80,
            "local_pdf": 65, "error": 50,
        }
        for i, f in enumerate(REPORT_FIELDS, 1):
            ws.column_dimensions[get_column_letter(i)].width = wide.get(f, 22)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ss = wb.create_sheet("Summary")
        total      = len(rows)
        downloaded = sum(1 for r in rows if r.get("local_pdf"))
        no_oa      = sum(1 for r in rows if r.get("oa_status") == "closed")
        failed_dl  = total - downloaded - no_oa

        ss.append(["Metric", "Count"])
        ss.append(["Total DOIs", total])
        ss.append(["Downloaded", downloaded])
        ss.append(["No OA version found (closed access)", no_oa])
        ss.append(["OA found but download failed", failed_dl])

        wb.save(xlsx_path)
        print(f"\nReport saved: {xlsx_path}")

    except ImportError:
        print(f"\nReport saved (CSV only): {csv_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    create_dirs()
    quarantine_broken()

    records = load_scopus()
    total   = len(records)

    print()
    print("=" * 68)
    print("  OPENALEX EXTENDED LEGAL OA DOWNLOADER")
    print(f"  {total} DOIs loaded from {SCOPUS_EXCEL.name}")
    print("=" * 68)

    rows = []

    for i, rec in enumerate(records, 1):
        doi      = rec["doi"]
        paper_id = rec["paper_id"]
        title    = rec["title"]

        pdf_name = f"{safe_filename(paper_id)}_{safe_filename(title)}.pdf"
        out_path = PDF_DIR / pdf_name

        print(f"\n[{i}/{total}] {doi}")
        print(f"  Paper ID : {paper_id}")
        print(f"  PDF name : {pdf_name}")

        row = {
            **rec,
            "oa_status":      "",
            "oa_url_count":   0,
            "pdf_url_tried":  "",
            "all_oa_urls":    "",
            "download_status": "",
            "local_pdf":      "",
            "error":          "",
        }

        if out_path.exists() and is_valid_pdf(out_path):
            print("  ✓ Already downloaded.")
            row["local_pdf"]       = str(out_path)
            row["download_status"] = "Already downloaded"
            rows.append(row)
            continue

        print("  Querying OpenAlex …")
        data, err = query_openalex(doi)

        if err and not data:
            print(f"  ✗ OpenAlex error: {err}")
            row["error"]           = err
            row["download_status"] = "OpenAlex API error"
            rows.append(row)
            continue

        if not data:
            print("  ✗ DOI not found in OpenAlex.")
            row["download_status"] = "DOI not found in OpenAlex"
            rows.append(row)
            continue

        oa_info   = data.get("open_access", {})
        oa_status = oa_info.get("oa_status", "unknown")
        row["oa_status"] = oa_status

        candidates = extract_pdf_candidates(data)
        row["oa_url_count"] = len(candidates)
        row["all_oa_urls"]  = " | ".join(c["url"] for c in candidates)

        print(f"  OA status     : {oa_status}")
        print(f"  PDF candidates: {len(candidates)}")

        if not candidates:
            msg = f"No OA PDF URLs found (oa_status={oa_status})"
            print(f"  ✗ {msg}")
            row["download_status"] = msg
            rows.append(row)
            continue

        downloaded = False

        for j, cand in enumerate(candidates, 1):
            url      = cand["url"]
            version  = cand["version"] or "unknown version"
            host     = cand["host_type"] or "unknown host"
            is_best  = "★ best_oa" if cand["is_best"] else f"candidate {j}"

            print(f"  [{is_best}] {host} / {version}")
            print(f"  URL: {url}")
            print("  Downloading …")

            row["pdf_url_tried"] = url

            ok, msg = download_pdf(url, out_path)

            if ok:
                print(f"  ✓ Saved → {out_path.name}")
                row["local_pdf"]       = str(out_path)
                row["download_status"] = f"Downloaded ({host} / {version})"
                downloaded = True
                break

            print(f"  ✗ {msg}")
            row["error"] = msg

        if not downloaded:
            row["download_status"] = (
                f"All {len(candidates)} OA URL(s) failed — see error column"
            )
            print(f"  ✗ All {len(candidates)} candidate(s) failed.")

        rows.append(row)
        time.sleep(0.5)   # be polite to OpenAlex

    write_report(rows, "04_openalex_report")

    dl_count = sum(1 for r in rows if r.get("local_pdf"))
    print()
    print("=" * 68)
    print(f"  Done.  Downloaded {dl_count}/{total}")
    print("=" * 68)
    print()


if __name__ == "__main__":
    main()
