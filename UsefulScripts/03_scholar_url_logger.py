#!/usr/bin/env python3
"""
Scholar URL Logger
==================
Searches Google Scholar by DOI for each paper in your Scopus Excel.
Does NOT attempt any downloads.

For each paper it logs:
  (a) Direct .pdf links found
  (b) MDPI / repository page links found
  (c) Any other download-related links found
  (d) The Scholar result page URL

Produces 4 output files (all in Research_Articles/Reports/):
  scholar_found_urls.csv       — all papers where Scholar found links
  scholar_not_found.csv        — papers with NO result on Scholar
  scholar_url_log.txt          — human-readable log, written live
  scholar_report.xlsx          — full Excel summary

Requirements
------------
    pip install playwright pandas openpyxl
    playwright install chromium

Usage
-----
    python 03_scholar_url_logger.py

Config (optional env vars)
--------------------------
    SCOPUS_EXCEL    path to xlsx    (default: scopus_articale_only.xlsx)
    SCOPUS_SHEET    sheet name      (default: Sheet1)
    SCHOLAR_DELAY   seconds between searches  (default: 8)
    SCHOLAR_HEADLESS  1 = headless  (default: 0 = visible, so you can solve CAPTCHAs)
"""

import csv
import os
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote, urlparse

# ── Config ────────────────────────────────────────────────────────────────────

SCOPUS_EXCEL     = Path(os.getenv("SCOPUS_EXCEL", "scopus_articale_only.xlsx"))
SCOPUS_SHEET     = os.getenv("SCOPUS_SHEET", "Sheet1")
SCHOLAR_DELAY    = int(os.getenv("SCHOLAR_DELAY", "8"))
SCHOLAR_HEADLESS = os.getenv("SCHOLAR_HEADLESS", "0") == "1"

ROOT       = Path("Research_Articles")
REPORT_DIR = ROOT / "Reports"

FOUND_CSV     = REPORT_DIR / "scholar_found_urls.csv"
NOT_FOUND_CSV = REPORT_DIR / "scholar_not_found.csv"
LOG_TXT       = REPORT_DIR / "scholar_url_log.txt"
REPORT_XLSX   = REPORT_DIR / "scholar_report.xlsx"

# Link classification
PDF_EXTENSIONS  = (".pdf",)
MDPI_DOMAINS    = ("mdpi.com",)
REPO_DOMAINS    = (
    "arxiv.org", "researchgate.net", "semanticscholar.org",
    "academia.edu", "biorxiv.org", "ssrn.com", "zenodo.org",
    "core.ac.uk", "base-search.net", "unpaywall.org",
    "europepmc.org", "ncbi.nlm.nih.gov", "hal.science",
    "figshare.com", "osf.io", "preprints.org",
)
SKIP_DOMAINS = (
    "/scholar",   # Scholar internal nav links
    "google.com/scholar",
    "accounts.google.com",
)

# CSV fields
FOUND_FIELDS = [
    "paper_id", "doi", "title", "year", "source_title",
    "scholar_result_title", "scholar_result_url",
    "direct_pdf_links",     # semicolon-separated .pdf URLs
    "mdpi_links",           # semicolon-separated mdpi.com URLs
    "repo_links",           # semicolon-separated known-repo URLs
    "other_links",          # other non-Scholar links found
    "all_links",            # everything combined
    "link_count",
]

NOT_FOUND_FIELDS = [
    "paper_id", "doi", "title", "year", "source_title", "reason",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def create_dirs():
    REPORT_DIR.mkdir(parents=True, exist_ok=True)


def clean_doi(v):
    v = str(v or "").strip().replace("\ufeff", "")
    v = re.sub(r"^\s*doi\s*:\s*", "", v, flags=re.I)
    v = re.sub(r"^https?://(dx\.)?doi\.org/", "", v, flags=re.I)
    return v.strip(" <>[](){}'\".,;")


def valid_doi(v):
    return bool(re.match(r"^10\.\d{4,9}/\S+$", v, flags=re.I))


def safe_filename(v, limit=180):
    v = str(v or "").strip()
    v = re.sub(r'[\\/:*?"<>|]', "_", v)
    v = re.sub(r"\s+", " ", v)
    return v[:limit].strip()


def normalize(v):
    v = str(v or "").lower()
    return re.sub(r"[^a-z0-9 ]+", " ", v).strip()


def title_similarity(a, b):
    import difflib
    a, b = normalize(a), normalize(b)
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def classify_url(href: str) -> str:
    """Return 'pdf', 'mdpi', 'repo', 'skip', or 'other'."""
    if not href:
        return "skip"
    low = href.lower()
    if any(s in low for s in SKIP_DOMAINS):
        return "skip"
    if low.endswith(".pdf") or "/pdf/" in low or "format=pdf" in low or "type=pdf" in low:
        return "pdf"
    domain = urlparse(href).netloc.lower()
    if any(d in domain for d in MDPI_DOMAINS):
        return "mdpi"
    if any(d in domain for d in REPO_DOMAINS):
        return "repo"
    return "other"


def log(msg: str):
    """Print and append to the live log file simultaneously."""
    print(msg)
    with open(LOG_TXT, "a", encoding="utf-8") as f:
        f.write(msg + "\n")


# ── Scopus loader ─────────────────────────────────────────────────────────────

def load_scopus():
    try:
        import pandas as pd
    except ImportError:
        print("ERROR: pip install pandas openpyxl")
        sys.exit(1)

    if not SCOPUS_EXCEL.exists():
        print(f"ERROR: {SCOPUS_EXCEL.resolve()} not found.")
        sys.exit(1)

    df = pd.read_excel(SCOPUS_EXCEL, sheet_name=SCOPUS_SHEET)
    df.columns = [str(c).strip() for c in df.columns]

    for col in ("DOI", "S.No"):
        if col not in df.columns:
            print(f"ERROR: Column '{col}' not found. Columns: {list(df.columns)}")
            sys.exit(1)

    records = []
    for _, row in df.iterrows():
        doi = clean_doi(row.get("DOI", ""))
        if not doi or not valid_doi(doi):
            continue
        s_no = str(row.get("S.No", "")).strip().rstrip(".0") or ""
        paper_id = s_no or str(row.get("index", "")).strip().rstrip(".0")
        if not paper_id:
            continue
        records.append({
            "paper_id":     paper_id,
            "doi":          doi,
            "title":        str(row.get("Title",        "")).strip(),
            "year":         str(row.get("Year",         "")).strip().rstrip(".0"),
            "source_title": str(row.get("Source title", "")).strip(),
        })
    return records


# ── Live CSV writers ───────────────────────────────────────────────────────────

def append_found(row: dict):
    is_new = not FOUND_CSV.exists()
    with open(FOUND_CSV, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=FOUND_FIELDS, extrasaction="ignore")
        if is_new:
            w.writeheader()
        w.writerow({k: row.get(k, "") for k in FOUND_FIELDS})


def append_not_found(row: dict):
    is_new = not NOT_FOUND_CSV.exists()
    with open(NOT_FOUND_CSV, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=NOT_FOUND_FIELDS, extrasaction="ignore")
        if is_new:
            w.writeheader()
        w.writerow({k: row.get(k, "") for k in NOT_FOUND_FIELDS})


# ── Excel report ───────────────────────────────────────────────────────────────

def write_excel(found_rows: list, not_found_rows: list):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("  (openpyxl not installed — skipping Excel report)")
        return

    wb = Workbook()

    # ── Found sheet ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Found - URLs"
    ws1.append(FOUND_FIELDS)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="C6EFCE")

    for row in found_rows:
        ws1.append([row.get(f, "") for f in FOUND_FIELDS])

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    wide = {
        "title": 55, "doi": 35, "scholar_result_title": 55,
        "scholar_result_url": 60,
        "direct_pdf_links": 70, "mdpi_links": 70,
        "repo_links": 70, "other_links": 70, "all_links": 80,
    }
    for i, f in enumerate(FOUND_FIELDS, 1):
        ws1.column_dimensions[get_column_letter(i)].width = wide.get(f, 18)
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Not found sheet ──────────────────────────────────────────
    ws2 = wb.create_sheet("Not Found on Scholar")
    ws2.append(NOT_FOUND_FIELDS)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFCCCC")

    for row in not_found_rows:
        ws2.append([row.get(f, "") for f in NOT_FOUND_FIELDS])

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    for i, f in enumerate(NOT_FOUND_FIELDS, 1):
        ws2.column_dimensions[get_column_letter(i)].width = (
            55 if f in ("title", "reason") else 35 if f == "doi" else 18
        )

    # ── Summary sheet ────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    total      = len(found_rows) + len(not_found_rows)
    with_pdf   = sum(1 for r in found_rows if r.get("direct_pdf_links"))
    with_mdpi  = sum(1 for r in found_rows if r.get("mdpi_links"))
    with_repo  = sum(1 for r in found_rows if r.get("repo_links"))
    with_other = sum(1 for r in found_rows if r.get("other_links"))
    not_found  = len(not_found_rows)

    ws3.append(["Metric", "Count"])
    ws3.append(["Total DOIs processed", total])
    ws3.append(["Found on Scholar (any link)", len(found_rows)])
    ws3.append(["  → with direct .pdf link", with_pdf])
    ws3.append(["  → with MDPI link", with_mdpi])
    ws3.append(["  → with repository link", with_repo])
    ws3.append(["  → with other link only", with_other])
    ws3.append(["Not found on Scholar", not_found])

    for cell in ws3["A"]:
        cell.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 12

    wb.save(REPORT_XLSX)
    log(f"\n  Excel report saved: {REPORT_XLSX}")


# ── Scholar search ─────────────────────────────────────────────────────────────

def scholar_search(page, doi: str, title: str) -> dict:
    """
    Search Scholar by DOI. Collect and classify all links from matching
    results. Returns a dict with keys:
        found, result_title, result_url,
        direct_pdf_links, mdpi_links, repo_links, other_links, all_links
    """
    result = {
        "found": False,
        "result_title": "",
        "result_url": "",
        "direct_pdf_links": [],
        "mdpi_links": [],
        "repo_links": [],
        "other_links": [],
        "all_links": [],
        "reason": "",
    }

    url = f"https://scholar.google.com/scholar?q={quote(doi, safe='')}"

    try:
        page.goto(url, timeout=60000, wait_until="domcontentloaded")
        page.wait_for_timeout(3000)
    except Exception as e:
        result["reason"] = f"Navigation error: {e}"
        return result

    body = page.inner_text("body").lower()

    # CAPTCHA check
    challenge_terms = [
        "unusual traffic", "captcha", "i'm not a robot",
        "verify you are human", "not a robot",
    ]
    if any(t in body for t in challenge_terms):
        if SCHOLAR_HEADLESS:
            result["reason"] = "CAPTCHA — run with SCHOLAR_HEADLESS=0"
            return result
        print()
        print("  *** CAPTCHA appeared — solve it in the browser, then press ENTER ***")
        input("  Press ENTER when Scholar results are visible … ")
        page.wait_for_timeout(2000)
        body = page.inner_text("body").lower()

    # Find result blocks — outer div contains both title (gs_ri) and
    # PDF badge (gs_ggs), so we scope to the outer wrapper
    blocks = page.locator("div.gs_r.gs_scl")
    if blocks.count() == 0:
        blocks = page.locator("div.gs_ri")

    if blocks.count() == 0:
        result["reason"] = "No results page rendered"
        return result

    for i in range(min(blocks.count(), 10)):
        block = blocks.nth(i)

        try:
            res_title = block.locator("h3.gs_rt").inner_text()
        except Exception:
            res_title = ""

        try:
            res_url = block.locator("h3.gs_rt a").get_attribute("href") or ""
        except Exception:
            res_url = ""

        score = title_similarity(title, res_title)

        if score < 0.40:
            continue

        result["found"]        = True
        result["result_title"] = res_title
        result["result_url"]   = res_url

        # Collect every link in this result block
        links = block.locator("a")
        for j in range(links.count()):
            try:
                href = links.nth(j).get_attribute("href") or ""
                text = links.nth(j).inner_text().strip()
            except Exception:
                continue

            if not href:
                continue

            kind = classify_url(href)
            if kind == "skip":
                continue

            if href not in result["all_links"]:
                result["all_links"].append(href)

            if kind == "pdf" and href not in result["direct_pdf_links"]:
                result["direct_pdf_links"].append(href)
            elif kind == "mdpi" and href not in result["mdpi_links"]:
                result["mdpi_links"].append(href)
            elif kind == "repo" and href not in result["repo_links"]:
                result["repo_links"].append(href)
            elif kind == "other" and href not in result["other_links"]:
                result["other_links"].append(href)

        # We matched — stop after the first good result
        break

    if not result["found"]:
        result["reason"] = "No matching result found (low title similarity)"

    return result


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    create_dirs()

    # Clear log file from previous runs
    LOG_TXT.write_text("", encoding="utf-8")

    records = load_scopus()
    total   = len(records)

    log("=" * 68)
    log("  SCHOLAR URL LOGGER")
    log(f"  {total} DOIs from {SCOPUS_EXCEL.name}")
    log(f"  Delay: {SCHOLAR_DELAY}s  |  Headless: {SCHOLAR_HEADLESS}")
    log(f"  Found URLs  → {FOUND_CSV}")
    log(f"  Not found   → {NOT_FOUND_CSV}")
    log(f"  Live log    → {LOG_TXT}")
    log("=" * 68)

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("ERROR: pip install playwright && playwright install chromium")
        sys.exit(1)

    found_rows     = []
    not_found_rows = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=SCHOLAR_HEADLESS,
            args=["--disable-extensions-except=",
                  "--disable-component-extensions-with-background-pages"],
        )
        context = browser.new_context(
            ignore_https_errors=True,
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/138.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()

        try:
            for i, rec in enumerate(records, 1):
                doi      = rec["doi"]
                paper_id = rec["paper_id"]
                title    = rec["title"]

                log(f"\n[{i}/{total}] {doi}")
                log(f"  Paper ID : {paper_id}")
                log(f"  Title    : {title[:80]}")

                res = scholar_search(page, doi, title)

                if res["found"]:
                    row = {
                        **rec,
                        "scholar_result_title": res["result_title"],
                        "scholar_result_url":   res["result_url"],
                        "direct_pdf_links":     " ; ".join(res["direct_pdf_links"]),
                        "mdpi_links":           " ; ".join(res["mdpi_links"]),
                        "repo_links":           " ; ".join(res["repo_links"]),
                        "other_links":          " ; ".join(res["other_links"]),
                        "all_links":            " ; ".join(res["all_links"]),
                        "link_count":           len(res["all_links"]),
                    }
                    found_rows.append(row)
                    append_found(row)

                    log(f"  ✓ Found: {res['result_title'][:70]}")
                    if res["direct_pdf_links"]:
                        log(f"  [PDF links]  {len(res['direct_pdf_links'])}:")
                        for u in res["direct_pdf_links"]:
                            log(f"    {u}")
                    if res["mdpi_links"]:
                        log(f"  [MDPI links] {len(res['mdpi_links'])}:")
                        for u in res["mdpi_links"]:
                            log(f"    {u}")
                    if res["repo_links"]:
                        log(f"  [Repo links] {len(res['repo_links'])}:")
                        for u in res["repo_links"]:
                            log(f"    {u}")
                    if res["other_links"]:
                        log(f"  [Other links] {len(res['other_links'])}:")
                        for u in res["other_links"]:
                            log(f"    {u}")
                    if not res["all_links"]:
                        log("  (no external links in this result block)")

                else:
                    row = {
                        **rec,
                        "reason": res.get("reason", "Not found"),
                    }
                    not_found_rows.append(row)
                    append_not_found(row)
                    log(f"  ✗ Not found — {res.get('reason', '')}")

                if i < total:
                    log(f"  Waiting {SCHOLAR_DELAY}s …")
                    time.sleep(SCHOLAR_DELAY)

        finally:
            try:
                browser.close()
            except Exception:
                pass

    write_excel(found_rows, not_found_rows)

    log("\n" + "=" * 68)
    log(f"  DONE")
    log(f"  Found on Scholar     : {len(found_rows)}/{total}")
    log(f"  Not found on Scholar : {len(not_found_rows)}/{total}")
    log(f"  Found CSV   → {FOUND_CSV}")
    log(f"  Not found   → {NOT_FOUND_CSV}")
    log(f"  Excel       → {REPORT_XLSX}")
    log("=" * 68)


if __name__ == "__main__":
    main()
