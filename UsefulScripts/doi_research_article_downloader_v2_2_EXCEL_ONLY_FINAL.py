#!/usr/bin/env python3
"""
DOI Research Article Downloader - Version 2.2
==============================================

MASTER INDEX / ONLY INPUT:
    Scopus Excel dataset

PAPER ID / FILE NUMBER:
    Scopus "S.No" column (preferred)
    Falls back to "index" only if S.No is unavailable.

RETRIEVAL ORDER:
    0. Existing local PDF
    1. Unpaywall
    2. Consensus
    3. Google Scholar
    4. Extended legal OA search (OpenAlex)
    5. Manual review

IMPORTANT:
- The Scopus Excel workbook is the ONLY input source for DOIs.
- No DOI text file is read or required.
- Existing PDFs are preserved.
- No paywalls, CAPTCHA, login, or access-control bypass is attempted.
- Google Scholar uses a visible Chromium browser.
- Scholar's right-side [PDF] / gs_ggs link is explicitly detected.
- The PDF is downloaded and verified BEFORE the browser is closed.
- Consensus uses its official API when CONSENSUS_API_KEY is configured.
- Without a Consensus API key, the script records a Consensus search URL.
- OpenAlex is used as a legal open-access fallback after Scholar.
"""

import csv
import difflib
import os
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote

import requests


# ============================================================
# CONFIGURATION
# ============================================================

# Your Scopus master Excel file.
SCOPUS_EXCEL = Path(
    os.getenv(
        "SCOPUS_EXCEL",
        "scopus_articale_only.xlsx"
    )
)

# Sheet containing the original Scopus-style records.
SCOPUS_SHEET = os.getenv(
    "SCOPUS_SHEET",
    "researchArticalOnly"
)

ROOT = Path("Research_Articles")
PDF_DIR = ROOT / "PDFs"
FAIL_DIR = ROOT / "Failed"
REPORT_DIR = ROOT / "Reports"

UNPAYWALL_EMAIL = os.getenv(
    "UNPAYWALL_EMAIL",
    "amanpalrayat@gmail.com"
)

CONSENSUS_API_KEY = os.getenv(
    "CONSENSUS_API_KEY",
    ""
)

ENABLE_GOOGLE_SCHOLAR = True
ENABLE_OPENALEX = True

# Delay between Scholar searches.
SCHOLAR_DELAY = 8

TIMEOUT = 30
MAX_RETRIES = 3

# Existing PDF matching threshold.
TITLE_MATCH_THRESHOLD = 0.92

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "ResearchArticleDownloader/2.2 "
        "(academic literature retrieval)"
    ),
    "Accept-Language": "en-US,en;q=0.9",
})


FIELDS = [
    "paper_id",
    "scopus_index",
    "scopus_s_no",
    "doi",
    "title",
    "authors",
    "year",
    "source_title",

    "existing_pdf",
    "existing_pdf_path",

    "unpaywall_status",
    "oa_status",
    "oa_host_type",
    "unpaywall_pdf_url",

    "consensus_status",
    "consensus_title",
    "consensus_url",
    "consensus_match_score",

    "scholar_status",
    "scholar_result_title",
    "scholar_result_url",
    "scholar_pdf_url",

    "openalex_status",
    "openalex_pdf_url",

    "local_pdf",
    "final_source",
    "download_status",
    "google_scholar_url",
    "consensus_search_url",
    "openalex_url",
    "error",
]


# ============================================================
# GENERAL HELPERS
# ============================================================

def create_directories():
    for folder in (
        ROOT,
        PDF_DIR,
        FAIL_DIR,
        REPORT_DIR,
    ):
        folder.mkdir(
            parents=True,
            exist_ok=True
        )


def clean_doi(value):
    value = str(value or "").strip()
    value = value.replace("\ufeff", "")

    value = re.sub(
        r"^\s*doi\s*:\s*",
        "",
        value,
        flags=re.I
    )

    value = re.sub(
        r"^https?://(dx\.)?doi\.org/",
        "",
        value,
        flags=re.I
    )

    return value.strip(
        " <>[](){}'\".,;"
    )


def valid_doi(value):
    return bool(
        re.match(
            r"^10\.\d{4,9}/\S+$",
            value,
            flags=re.I
        )
    )


def normalize_text(value):
    value = str(value or "").lower()
    value = re.sub(
        r"[^a-z0-9]+",
        " ",
        value
    )

    return re.sub(
        r"\s+",
        " ",
        value
    ).strip()


def safe_filename(value, limit=180):
    value = str(value or "untitled")

    value = re.sub(
        r'[<>:"/\\|?*\x00-\x1f]',
        "_",
        value
    )

    value = re.sub(
        r"\s+",
        " ",
        value
    ).strip().rstrip(". ")

    return (
        value or "untitled"
    )[:limit]


def similarity(a, b):
    return difflib.SequenceMatcher(
        None,
        normalize_text(a)[:500],
        normalize_text(b)[:500]
    ).ratio()


def scholar_url(doi, title):
    query = title or doi

    return (
        "https://scholar.google.com/scholar?q="
        + quote(query)
    )


def consensus_search_url(doi):
    return (
        "https://consensus.app/search/?q="
        + quote(doi)
    )


def openalex_url(doi):
    return (
        "https://api.openalex.org/works/https://doi.org/"
        + quote(doi, safe="")
    )


def get_json(
    url,
    params=None,
    headers=None
):
    last_error = ""

    for attempt in range(
        MAX_RETRIES
    ):

        try:

            response = SESSION.get(
                url,
                params=params,
                headers=headers,
                timeout=TIMEOUT
            )

            if response.status_code == 200:

                return (
                    response.json(),
                    ""
                )

            last_error = (
                f"HTTP {response.status_code}"
            )

        except Exception as exc:

            last_error = str(exc)

        if attempt < MAX_RETRIES - 1:

            time.sleep(
                2 * (attempt + 1)
            )

    return {}, last_error


# ============================================================
# SCOPUS MASTER EXCEL
# ============================================================

def load_scopus_master():
    """
    Loads the Scopus Excel master mapping.

    Preferred columns:
        S.No
        DOI
        Title
        Year
        Authors
        Source title

    The S.No is used as the permanent paper/file number.
    """

    try:
        import pandas as pd
    except ImportError:
        print(
            "ERROR: pandas is required."
        )
        sys.exit(1)

    if not SCOPUS_EXCEL.exists():

        print()
        print(
            "ERROR: Scopus Excel file not found:"
        )
        print(
            SCOPUS_EXCEL.resolve()
        )
        print()
        print(
            "Set SCOPUS_EXCEL if your file has a "
            "different name/path."
        )

        sys.exit(1)

    try:

        dataframe = pd.read_excel(
            SCOPUS_EXCEL,
            sheet_name=SCOPUS_SHEET
        )

    except Exception as exc:

        print()
        print(
            "ERROR reading Scopus Excel:"
        )
        print(exc)
        sys.exit(1)

    # Normalize column names.
    dataframe.columns = [
        str(column).strip()
        for column in dataframe.columns
    ]

    required = [
        "DOI",
        "S.No",
    ]

    missing = [
        column
        for column in required
        if column not in dataframe.columns
    ]

    if missing:

        print()
        print(
            "ERROR: Required columns missing:"
        )
        print(
            ", ".join(missing)
        )
        print()
        print(
            "Found columns:"
        )
        print(
            list(dataframe.columns)
        )

        sys.exit(1)

    records = {}

    for row_number, (_, row) in enumerate(dataframe.iterrows(), start=1):

        doi = clean_doi(
            row.get("DOI", "")
        )

        if not doi:
            continue

        # S.No is the permanent paper ID.
        s_no = row.get(
            "S.No",
            ""
        )

        index_value = row.get(
            "index",
            ""
        )

        # Convert Excel numeric values cleanly.
        def clean_number(value):

            if value is None:
                return ""

            text = str(value).strip()

            if text.endswith(".0"):
                text = text[:-2]

            return text

        s_no = clean_number(
            s_no
        )

        index_value = clean_number(
            index_value
        )

        if s_no:
            paper_id = s_no
        elif index_value:
            paper_id = index_value
        else:
            continue

        record = {
            "paper_id":
                paper_id,

            "scopus_index":
                index_value,

            "scopus_s_no":
                s_no,

            "doi":
                doi,

            "title":
                str(
                    row.get(
                        "Title",
                        ""
                    )
                ).strip(),

            "authors":
                str(
                    row.get(
                        "Authors",
                        ""
                    )
                ).strip(),

            "year":
                clean_number(
                    row.get(
                        "Year",
                        ""
                    )
                ),

            "source_title":
                str(
                    row.get(
                        "Source title",
                        ""
                    )
                ).strip(),
        }

        # IMPORTANT:
        # Keep EVERY Excel row.  Do NOT deduplicate by DOI.
        # S.No / paper_id is the permanent identity of the paper.
        # This preserves all Scopus rows even when two rows share
        # the same DOI.
        record_key = (
            f"{paper_id}__{doi.lower()}"
            if paper_id
            else f"row_{row_number}__{doi.lower()}"
        )

        records[record_key] = record

    print()
    print(
        f"Scopus master records loaded: "
        f"{len(records)}"
    )

    return records


# ============================================================
# DOI INPUT — SCOPUS EXCEL ONLY
# ============================================================

def load_doi_list(scopus_records):
    """
    Build the processing list exclusively from the Scopus Excel.

    IMPORTANT:
    - No DOI .txt file is read.
    - Every Excel row with a valid DOI is retained.
    - DOIs are NOT deduplicated.
    - Scopus S.No is the permanent paper/file ID.
    - Processing follows ascending Scopus S.No.
    """

    records = list(scopus_records.values())

    def sort_key(record):
        value = str(record.get("scopus_s_no", "")).strip()

        try:
            return (0, float(value))
        except (TypeError, ValueError):
            return (1, value)

    records.sort(key=sort_key)

    return records


# ============================================================
# PREVIOUS REPORT / EXISTING PDF PROTECTION
# ============================================================

def load_previous_rows():

    candidates = [
        REPORT_DIR /
        "master_paper_database_progress.csv",

        REPORT_DIR /
        "master_paper_database.csv",
    ]

    for report in candidates:

        if not report.exists():
            continue

        try:

            with open(
                report,
                encoding="utf-8-sig",
                newline=""
            ) as file:

                reader = csv.DictReader(
                    file
                )

                result = {}

                for row in reader:

                    doi = clean_doi(
                        row.get(
                            "doi",
                            ""
                        )
                    )

                    if doi:
                        result[
                            doi.lower()
                        ] = row

                if result:

                    print(
                        "Previous progress loaded:"
                    )

                    print(
                        report
                    )

                return result

        except Exception:
            pass

    return {}


def find_existing_pdf(
    paper_id,
    doi,
    title,
    previous
):

    old = previous.get(
        doi.lower(),
        {}
    )

    old_path = old.get(
        "local_pdf",
        ""
    )

    if (
        old_path
        and Path(old_path).exists()
    ):

        return (
            Path(old_path),
            "previous report"
        )

    # Prefer a PDF that already contains the Scopus
    # paper ID in its filename.
    if paper_id:

        pattern = (
            f"{safe_filename(paper_id)}_*.pdf"
        )

        matches = list(
            PDF_DIR.glob(pattern)
        )

        if matches:

            return (
                matches[0],
                "Scopus paper ID filename"
            )

    if not title:
        return None, ""

    best = None
    best_score = 0

    if PDF_DIR.exists():

        for pdf in PDF_DIR.glob(
            "*.pdf"
        ):

            score = similarity(
                title,
                pdf.stem
            )

            if score > best_score:

                best_score = score
                best = pdf

    if (
        best
        and best_score >= TITLE_MATCH_THRESHOLD
    ):

        return (
            best,
            f"title match {best_score:.2f}"
        )

    return None, ""


# ============================================================
# CROSSREF
# ============================================================

def get_crossref_metadata(doi):

    url = (
        "https://api.crossref.org/works/"
        + quote(
            doi,
            safe=""
        )
    )

    data, error = get_json(
        url
    )

    if not data:
        return {}, error

    message = data.get(
        "message",
        {}
    )

    authors = []

    for author in message.get(
        "author",
        []
    ):

        name = " ".join(
            x for x in (
                author.get(
                    "given",
                    ""
                ),
                author.get(
                    "family",
                    ""
                )
            )
            if x
        )

        if name:
            authors.append(name)

    published = (
        message.get(
            "published-print"
        )
        or
        message.get(
            "published-online"
        )
        or
        message.get(
            "issued"
        )
        or
        {}
    )

    date_parts = published.get(
        "date-parts",
        [[]]
    )

    year = ""

    if (
        date_parts
        and date_parts[0]
    ):
        year = date_parts[0][0]

    abstract = message.get(
        "abstract",
        ""
    )

    abstract = re.sub(
        r"<[^>]+>",
        " ",
        abstract
    )

    abstract = re.sub(
        r"\s+",
        " ",
        abstract
    ).strip()

    return {
        "title":
            (
                message.get(
                    "title",
                    [""]
                )[0]
                if message.get("title")
                else ""
            ),

        "authors":
            "; ".join(
                authors
            ),

        "year":
            year,

        "source_title":
            (
                message.get(
                    "container-title",
                    [""]
                )[0]
                if message.get(
                    "container-title"
                )
                else ""
            ),

        "abstract":
            abstract,
    }, ""


# ============================================================
# UNPAYWALL - STAGE 1
# ============================================================

def get_unpaywall(doi):

    if "@" not in UNPAYWALL_EMAIL:

        return (
            {},
            "UNPAYWALL_EMAIL is not configured"
        )

    url = (
        "https://api.unpaywall.org/v2/"
        + quote(
            doi,
            safe=""
        )
    )

    return get_json(
        url,
        params={
            "email":
                UNPAYWALL_EMAIL
        }
    )


def unpaywall_pdf_candidates(data):

    candidates = []

    locations = []

    best = data.get(
        "best_oa_location"
    )

    if best:
        locations.append(
            best
        )

    locations.extend(
        data.get(
            "oa_locations",
            []
        ) or []
    )

    seen = set()

    for location in locations:

        if not isinstance(
            location,
            dict
        ):
            continue

        url = location.get(
            "url_for_pdf"
        )

        if (
            url
            and url not in seen
        ):

            candidates.append(
                url
            )

            seen.add(
                url
            )

    return candidates


# ============================================================
# CONSENSUS - STAGE 2
# ============================================================

def consensus_search(
    doi,
    title
):

    if not CONSENSUS_API_KEY:

        return (
            {},
            "Consensus API key not configured"
        )

    headers = {
        "x-api-key":
            CONSENSUS_API_KEY,

        "Accept":
            "application/json",

        "User-Agent":
            SESSION.headers[
                "User-Agent"
            ],
    }

    # DOI is the strongest query.
    queries = [
        doi
    ]

    if title:
        queries.append(
            f'"{title}"'
        )

    last_error = ""

    for query in queries:

        data, error = get_json(
            "https://api.consensus.app/v1/quick_search",
            params={
                "query":
                    query,
                "page":
                    0,
            },
            headers=headers
        )

        if data:

            return (
                data,
                ""
            )

        last_error = error

    return (
        {},
        last_error
    )


def consensus_results(data):

    if not data:
        return []

    results = (
        data.get(
            "papers"
        )
        or
        data.get(
            "results"
        )
        or
        data.get(
            "data"
        )
        or
        []
    )

    if isinstance(
        results,
        dict
    ):
        results = [
            results
        ]

    return [
        item
        for item in results
        if isinstance(
            item,
            dict
        )
    ]


def find_best_consensus(
    results,
    title
):

    if not results:
        return None, 0

    best = None
    best_score = 0

    for item in results[:20]:

        candidate_title = item.get(
            "title",
            ""
        )

        score = similarity(
            title,
            candidate_title
        )

        if score > best_score:

            best = item
            best_score = score

    return (
        best,
        best_score
    )


def consensus_public_pdf_urls(
    result
):

    urls = []

    # These are optional fields some API responses/integrations
    # may expose. The official Consensus search response primarily
    # provides a paper URL rather than a PDF URL.
    for key in (
        "pdf_url",
        "full_text_url",
        "open_access_url",
    ):

        value = result.get(
            key
        )

        if (
            isinstance(
                value,
                str
            )
            and value.startswith(
                (
                    "http://",
                    "https://"
                )
            )
        ):

            if value not in urls:
                urls.append(
                    value
                )

    return urls


# ============================================================
# GOOGLE SCHOLAR - STAGE 3
# ============================================================

def scholar_find_and_download(
    doi,
    title,
    output_path
):

    result = {
        "status":
            "No Scholar result",

        "found":
            False,

        "result_title":
            "",

        "result_url":
            "",

        "pdf_url":
            "",
    }

    try:

        from playwright.sync_api import (
            sync_playwright
        )

    except ImportError:

        result["status"] = (
            "Playwright not installed"
        )

        return result

    with sync_playwright() as playwright:

        browser = playwright.chromium.launch(
            headless=False
        )

        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 "
                "(Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 "
                "(KHTML, like Gecko) "
                "Chrome/138.0.0.0 "
                "Safari/537.36"
            )
        )

        page = context.new_page()

        try:

            page.goto(
                scholar_url(
                    doi,
                    title
                ),
                wait_until="domcontentloaded",
                timeout=60000
            )

            time.sleep(2)

            body_text = (
                page.locator(
                    "body"
                )
                .inner_text()
                .lower()
            )

            challenge_terms = [
                "captcha",
                "recaptcha",
                "unusual traffic",
                "not a robot",
                "our systems have detected",
            ]

            if any(
                term in body_text
                for term in challenge_terms
            ):

                print()
                print(
                    "Google Scholar verification "
                    "appeared."
                )

                print(
                    "Solve it manually in the "
                    "browser window."
                )

                input(
                    "Press ENTER after Scholar "
                    "is accessible..."
                )

            results = page.locator(
                "div.gs_ri"
            )

            count = results.count()

            if count == 0:

                result["status"] = (
                    "No Scholar results"
                )

                return result

            for i in range(
                min(
                    count,
                    10
                )
            ):

                item = results.nth(
                    i
                )

                try:

                    result_title = (
                        item.locator(
                            "h3.gs_rt"
                        )
                        .inner_text()
                    )

                except Exception:

                    result_title = ""

                try:

                    result_url = (
                        item.locator(
                            "h3.gs_rt a"
                        )
                        .get_attribute(
                            "href"
                        )
                    ) or ""

                except Exception:

                    result_url = ""

                score = similarity(
                    title,
                    result_title
                )

                # ------------------------------------------------
                # CRITICAL V2.2 FIX:
                # Google Scholar's right-side [PDF] link is
                # commonly inside div.gs_ggs.
                # ------------------------------------------------

                pdf_url = ""

                # IMPORTANT: Google Scholar has used multiple
                # DOM variants for the right-side document link.
                # Examples include:
                #   div.gs_ggs
                #   div.gs_or_ggsm
                # The latter is the structure currently seen for
                # links such as:
                #   <div class="gs_or_ggsm gs_press">
                #       <a href="https://.../paper.pdf">
                #           <span class="gs_ctg2">[PDF]</span>
                #       </a>
                #   </div>
                #
                # Search both containers and explicitly prefer
                # links containing the [PDF] marker.

                pdf_containers = item.locator(
                    "div.gs_or_ggsm, div.gs_ggs"
                )

                if pdf_containers.count() > 0:

                    pdf_links = (
                        pdf_containers.locator(
                            "a[href]"
                        )
                    )

                    # First pass: links whose visible text or
                    # child span explicitly says [PDF].
                    pdf_found = False

                    for j in range(
                        pdf_links.count()
                    ):
                        link = pdf_links.nth(j)

                        try:
                            href = (
                                link.get_attribute("href")
                                or ""
                            )
                            link_text = (
                                link.inner_text()
                                or ""
                            )
                        except Exception:
                            continue

                        if (
                            href
                            and
                            (
                                "[pdf]"
                                in link_text.lower()
                                or
                                link.locator(
                                    "span.gs_ctg2"
                                ).count() > 0
                            )
                        ):
                            pdf_url = href
                            pdf_found = True
                            break

                    # Second pass: any href in the Scholar
                    # right-side PDF container.
                    if not pdf_found:
                        for j in range(
                            pdf_links.count()
                        ):
                            href = (
                                pdf_links.nth(j)
                                .get_attribute("href")
                            )

                            if href:
                                pdf_url = href
                                break

                # ------------------------------------------------
                # Fallback: search the entire Scholar result for
                # the explicit [PDF] marker. This handles future
                # Scholar DOM changes.
                # ------------------------------------------------

                if not pdf_url:
                    try:
                        pdf_markers = item.locator(
                            "span.gs_ctg2"
                        )

                        for j in range(
                            pdf_markers.count()
                        ):
                            marker = pdf_markers.nth(j)
                            parent_link = marker.locator(
                                "xpath=ancestor::a[1]"
                            )

                            if parent_link.count():
                                href = (
                                    parent_link.first
                                    .get_attribute("href")
                                )

                                if href:
                                    pdf_url = href
                                    break

                    except Exception:
                        pass

                        href = (
                            pdf_links.nth(j)
                            .get_attribute(
                                "href"
                            )
                        )

                        if href:
                            pdf_url = href
                            break

                # ------------------------------------------------
                # Fallback 1: any PDF-looking link inside result.
                # ------------------------------------------------

                if not pdf_url:

                    links = item.locator(
                        "a[href]"
                    )

                    for j in range(
                        links.count()
                    ):

                        link = links.nth(j)

                        try:

                            href = (
                                link.get_attribute(
                                    "href"
                                )
                                or ""
                            )

                            text = (
                                link.inner_text()
                                or ""
                            )

                        except Exception:

                            continue

                        if (
                            ".pdf"
                            in href.lower()
                            or
                            "pdf"
                            in text.lower()
                        ):

                            pdf_url = href
                            break

                # ------------------------------------------------
                # Fallback 2: Scholar sometimes uses an external
                # link without .pdf in its URL but with [PDF].
                # ------------------------------------------------

                if not pdf_url:

                    try:

                        candidates = (
                            item.locator(
                                "div.gs_ggs a"
                            )
                        )

                        if candidates.count():

                            pdf_url = (
                                candidates
                                .first
                                .get_attribute(
                                    "href"
                                )
                                or ""
                            )

                    except Exception:
                        pass

                # ------------------------------------------------
                # Only accept a reasonably matching result.
                # If the first result has a PDF, a title match
                # above 0.65 is acceptable because DOI searches
                # often produce exact/near-exact first results.
                # ------------------------------------------------

                acceptable = (
                    score >= 0.65
                    or
                    (
                        i == 0
                        and bool(pdf_url)
                    )
                )

                if not acceptable:
                    continue

                result.update({
                    "status":
                        "Scholar result found",

                    "found":
                        True,

                    "result_title":
                        result_title,

                    "result_url":
                        result_url,

                    "pdf_url":
                        pdf_url,
                })

                print(
                    f"  Scholar result "
                    f"{i + 1}: "
                    f"{result_title}"
                )

                print(
                    f"  Title match: "
                    f"{score:.2f}"
                )

                if pdf_url:

                    print(
                        "  [PDF] link detected:"
                    )

                    print(
                        f"  {pdf_url}"
                    )

                    # --------------------------------------------
                    # IMPORTANT:
                    # Download while browser is still OPEN.
                    # --------------------------------------------

                    ok, message, final_url = (
                        download_pdf(
                            pdf_url,
                            output_path
                        )
                    )

                    if ok:

                        result["status"] = (
                            "Downloaded from Scholar"
                        )

                        print(
                            "  ✓ Scholar PDF downloaded "
                            "before browser closed."
                        )

                        return result

                    print(
                        "  Scholar PDF download failed "
                        "with direct HTTP request:"
                    )

                    print(
                        f"  {message}"
                    )

                    # Some publishers (including IEEE-hosted
                    # documents) may reject a plain requests.Session
                    # even though the PDF link is publicly reachable
                    # from the Scholar browser. Retry through the
                    # Playwright browser context before giving up.
                    try:
                        browser_response = context.request.get(
                            pdf_url,
                            timeout=60000,
                            headers={
                                "Referer": (
                                    "https://scholar.google.com/"
                                ),
                                "Accept": (
                                    "application/pdf,"
                                    "application/octet-stream,"
                                    "*/*"
                                ),
                            },
                        )

                        content_type = (
                            browser_response.headers.get(
                                "content-type",
                                ""
                            ).lower()
                        )

                        body = browser_response.body()

                        if (
                            browser_response.ok
                            and
                            (
                                body.startswith(b"%PDF-")
                                or
                                "application/pdf"
                                in content_type
                            )
                        ):
                            output_path.parent.mkdir(
                                parents=True,
                                exist_ok=True
                            )

                            with open(
                                output_path,
                                "wb"
                            ) as fh:
                                fh.write(body)

                            result["status"] = (
                                "Downloaded from Scholar"
                            )

                            print(
                                "  ✓ Scholar PDF downloaded "
                                "using browser-context request."
                            )

                            return result

                        print(
                            "  Browser-context PDF retry "
                            "did not return a valid PDF "
                            f"(HTTP {browser_response.status})."
                        )

                    except Exception as browser_exc:
                        print(
                            "  Browser-context PDF retry failed:"
                        )
                        print(
                            f"  {browser_exc}"
                        )

                else:

                    print(
                        "  No [PDF] link detected "
                        "for this Scholar result."
                    )

                # Continue checking another matching result
                # if the first one did not yield a PDF.

            if result["found"]:

                if result.get("pdf_url"):
                    result["status"] = (
                        "Scholar PDF link found - "
                        "download failed"
                    )
                else:
                    result["status"] = (
                        "Scholar result found - "
                        "no downloadable PDF"
                    )

        except Exception as exc:

            result["status"] = (
                f"Scholar error: {exc}"
            )

        finally:

            context.close()
            browser.close()

    return result


# ============================================================
# EXTENDED LEGAL OA SEARCH - OPENALEX
# ============================================================

def openalex_search(
    doi
):

    data, error = get_json(
        openalex_url(
            doi
        )
    )

    if not data:
        return (
            {},
            error
        )

    return (
        data,
        ""
    )


def openalex_pdf_candidates(
    data
):

    candidates = []

    locations = (
        data.get(
            "locations",
            []
        )
        or []
    )

    best = data.get(
        "best_oa_location"
    )

    ordered = []

    if best:
        ordered.append(
            best
        )

    ordered.extend(
        locations
    )

    seen = set()

    for location in ordered:

        if not isinstance(
            location,
            dict
        ):
            continue

        pdf = (
            location.get(
                "pdf_url"
            )
            or
            (
                location.get(
                    "pdf"
                ) or {}
            ).get(
                "url"
            )
        )

        if (
            pdf
            and pdf not in seen
        ):

            candidates.append(
                pdf
            )

            seen.add(
                pdf
            )

    return candidates


# ============================================================
# PDF DOWNLOAD + VERIFICATION
# ============================================================

def download_pdf(
    url,
    output_path
):

    try:

        response = SESSION.get(
            url,
            timeout=TIMEOUT,
            stream=True,
            allow_redirects=True
        )

        if response.status_code != 200:

            response.close()

            return (
                False,
                f"HTTP {response.status_code}",
                response.url
            )

        content_type = (
            response.headers.get(
                "Content-Type",
                ""
            ).lower()
        )

        # Read enough bytes to verify PDF magic header.
        first = response.raw.read(
            5
        )

        is_pdf = (
            first == b"%PDF-"
            or
            "application/pdf"
            in content_type
        )

        if not is_pdf:

            response.close()

            return (
                False,
                (
                    "Response is not a PDF "
                    f"(Content-Type: {content_type})"
                ),
                response.url
            )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        with open(
            output_path,
            "wb"
        ) as file:

            file.write(
                first
            )

            for chunk in response.iter_content(
                65536
            ):

                if chunk:
                    file.write(
                        chunk
                    )

        response.close()

        if (
            not output_path.exists()
            or
            output_path.stat().st_size
            < 1000
        ):

            output_path.unlink(
                missing_ok=True
            )

            return (
                False,
                "Suspiciously small PDF",
                response.url
            )

        return (
            True,
            "Downloaded",
            response.url
        )

    except Exception as exc:

        output_path.unlink(
            missing_ok=True
        )

        return (
            False,
            str(exc),
            url
        )


# ============================================================
# REPORTS
# ============================================================

def write_csv(
    rows,
    path
):

    with open(
        path,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as file:

        writer = csv.DictWriter(
            file,
            fieldnames=FIELDS,
            extrasaction="ignore"
        )

        writer.writeheader()

        for row in rows:

            writer.writerow({
                field:
                    row.get(
                        field,
                        ""
                    )
                for field in FIELDS
            })


def write_excel(
    rows,
    path
):

    try:

        from openpyxl import (
            Workbook
        )

        from openpyxl.styles import (
            Font,
            PatternFill,
            Alignment
        )

        from openpyxl.utils import (
            get_column_letter
        )

    except ImportError:

        return False

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = (
        "Master Database"
    )

    sheet.append(
        FIELDS
    )

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )

    for row in rows:

        sheet.append([
            row.get(
                field,
                ""
            )
            for field in FIELDS
        ])

    sheet.freeze_panes = (
        "A2"
    )

    sheet.auto_filter.ref = (
        sheet.dimensions
    )

    widths = {
        "title": 60,
        "authors": 40,
        "source_title": 40,
        "doi": 38,
        "unpaywall_pdf_url": 65,
        "consensus_url": 65,
        "scholar_result_url": 65,
        "scholar_pdf_url": 65,
        "openalex_pdf_url": 65,
        "local_pdf": 65,
        "error": 55,
    }

    for index, field in enumerate(
        FIELDS,
        start=1
    ):

        letter = get_column_letter(
            index
        )

        sheet.column_dimensions[
            letter
        ].width = widths.get(
            field,
            20
        )

    for row in sheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    summary = workbook.create_sheet(
        "Summary"
    )

    total = len(rows)

    existing = sum(
        r.get(
            "existing_pdf"
        ) == "Yes"
        for r in rows
    )

    downloaded = sum(
        r.get(
            "download_status",
            ""
        ).startswith(
            "Downloaded"
        )
        for r in rows
    )

    unpaywall = sum(
        r.get(
            "final_source"
        ) == "Unpaywall"
        for r in rows
    )

    consensus = sum(
        r.get(
            "final_source"
        ) == "Consensus"
        for r in rows
    )

    scholar = sum(
        r.get(
            "final_source"
        ) == "Google Scholar"
        for r in rows
    )

    openalex = sum(
        r.get(
            "final_source"
        ) == "OpenAlex"
        for r in rows
    )

    unresolved = sum(
        r.get(
            "local_pdf",
            ""
        ) == ""
        for r in rows
    )

    summary_rows = [
        (
            "Total DOI records",
            total
        ),
        (
            "Existing PDFs preserved",
            existing
        ),
        (
            "New PDFs downloaded",
            downloaded
        ),
        (
            "Downloaded via Unpaywall",
            unpaywall
        ),
        (
            "Downloaded via Consensus",
            consensus
        ),
        (
            "Downloaded via Google Scholar",
            scholar
        ),
        (
            "Downloaded via OpenAlex",
            openalex
        ),
        (
            "Unresolved / manual review",
            unresolved
        ),
    ]

    for item in summary_rows:

        summary.append(
            item
        )

    workbook.save(
        path
    )

    return True


# ============================================================
# MAIN
# ============================================================

def main():

    create_directories()

    print()
    print("=" * 72)
    print(
        "DOI RESEARCH ARTICLE DOWNLOADER V2.2"
    )
    print("=" * 72)

    print()
    print(
        "MASTER INDEX / ONLY INPUT:"
    )

    print(
        f"  {SCOPUS_EXCEL}"
    )

    print(
        "  DOI source: Scopus Excel only"
    )

    print(
        f"SCOPUS SHEET:"
    )

    print(
        f"  {SCOPUS_SHEET}"
    )

    print(
        "  Every Excel DOI row is retained; DOI duplicates are NOT removed."
    )

    print()
    print(
        "RETRIEVAL ORDER:"
    )

    print(
        "  0. Existing PDF"
    )

    print(
        "  1. Unpaywall"
    )

    print(
        "  2. Consensus"
    )

    print(
        "  3. Google Scholar"
    )

    print(
        "  4. OpenAlex legal OA"
    )

    print(
        "  5. Manual review"
    )

    # --------------------------------------------------------
    # Load master Scopus mapping.
    # --------------------------------------------------------

    scopus = (
        load_scopus_master()
    )

    # --------------------------------------------------------
    # DOI INPUT = SCOPUS EXCEL ONLY
    # --------------------------------------------------------

    dois = load_doi_list(
        scopus
    )

    print()
    print(
        "INPUT SOURCE:"
    )
    print(
        "  Scopus Excel only"
    )
    print(
        "  No DOI text file is used; every Excel DOI row is retained."
    )
    print(
        f"Excel records / DOIs to process: {len(dois)}"
    )

    previous = (
        load_previous_rows()
    )

    rows = []

    unmatched_scopus = []

    # ========================================================
    # PROCESS
    # ========================================================

    for number, master in enumerate(
        dois,
        start=1
    ):

        doi = master["doi"]

        print()
        print("-" * 72)
        print(
            f"[{number}/{len(dois)}] {doi}"
        )


        paper_id = (
            master["paper_id"]
        )

        title = (
            master["title"]
        )

        # ----------------------------------------------------
        # Crossref can improve incomplete metadata.
        # ----------------------------------------------------

        crossref, crossref_error = (
            get_crossref_metadata(
                doi
            )
        )

        if crossref:

            title = (
                title
                or
                crossref.get(
                    "title",
                    ""
                )
            )

        row = {
            field: ""
            for field in FIELDS
        }

        row.update({
            "paper_id":
                paper_id,

            "scopus_index":
                master[
                    "scopus_index"
                ],

            "scopus_s_no":
                master[
                    "scopus_s_no"
                ],

            "doi":
                doi,

            "title":
                title,

            "authors":
                master[
                    "authors"
                ]
                or
                crossref.get(
                    "authors",
                    ""
                ),

            "year":
                master[
                    "year"
                ]
                or
                crossref.get(
                    "year",
                    ""
                ),

            "source_title":
                master[
                    "source_title"
                ]
                or
                crossref.get(
                    "source_title",
                    ""
                ),

            "google_scholar_url":
                scholar_url(
                    doi,
                    title
                ),

            "consensus_search_url":
                consensus_search_url(
                    doi
                ),

            "openalex_url":
                openalex_url(
                    doi
                ),
        })

        if crossref_error:

            row["error"] = (
                "Crossref: "
                + crossref_error
            )

        # ----------------------------------------------------
        # PERMANENT FILENAME
        # ----------------------------------------------------

        filename = (
            f"{safe_filename(paper_id)}_"
            f"{safe_filename(title or doi)}.pdf"
        )

        output_path = (
            PDF_DIR
            /
            filename
        )

        print(
            f"  Paper ID: {paper_id}"
        )

        print(
            f"  PDF name: {filename}"
        )

        # ====================================================
        # EXISTING PDF
        # ====================================================

        existing, reason = (
            find_existing_pdf(
                paper_id,
                doi,
                title,
                previous
            )
        )

        if existing:

            row.update({
                "existing_pdf":
                    "Yes",

                "existing_pdf_path":
                    str(existing),

                "local_pdf":
                    str(existing),

                "final_source":
                    "Existing",

                "download_status":
                    "Existing - preserved",
            })

            print(
                "  ✓ Existing PDF preserved:"
            )

            print(
                f"    {existing}"
            )

            rows.append(
                row
            )

            write_csv(
                rows,
                REPORT_DIR /
                "master_paper_database_progress.csv"
            )

            continue

        # ====================================================
        # STAGE 1: UNPAYWALL
        # ====================================================

        print()
        print(
            "  [1/4] Unpaywall..."
        )

        upw, upw_error = (
            get_unpaywall(
                doi
            )
        )

        if upw:

            row[
                "unpaywall_status"
            ] = "Record found"

            row[
                "oa_status"
            ] = upw.get(
                "oa_status",
                ""
            )

            best = (
                upw.get(
                    "best_oa_location"
                )
                or
                {}
            )

            row[
                "oa_host_type"
            ] = best.get(
                "host_type",
                ""
            )

            for pdf_url in (
                unpaywall_pdf_candidates(
                    upw
                )
            ):

                ok, message, final_url = (
                    download_pdf(
                        pdf_url,
                        output_path
                    )
                )

                if ok:

                    row.update({
                        "unpaywall_pdf_url":
                            pdf_url,

                        "local_pdf":
                            str(output_path),

                        "final_source":
                            "Unpaywall",

                        "download_status":
                            "Downloaded - Unpaywall",
                    })

                    print(
                        "  ✓ PDF downloaded "
                        "from Unpaywall."
                    )

                    break

                row["error"] = (
                    f"{row['error']} | "
                    f"Unpaywall PDF: "
                    f"{message}"
                ).strip(
                    " |"
                )

        else:

            row[
                "unpaywall_status"
            ] = "No record"

            if upw_error:

                row["error"] = (
                    f"{row['error']} | "
                    f"Unpaywall: "
                    f"{upw_error}"
                ).strip(
                    " |"
                )

        if row["local_pdf"]:

            rows.append(
                row
            )

            write_csv(
                rows,
                REPORT_DIR /
                "master_paper_database_progress.csv"
            )

            continue

        # ====================================================
        # STAGE 2: CONSENSUS
        # ====================================================

        print()
        print(
            "  [2/4] Consensus..."
        )

        consensus_data, consensus_error = (
            consensus_search(
                doi,
                title
            )
        )

        if consensus_data:

            results = (
                consensus_results(
                    consensus_data
                )
            )

            best, score = (
                find_best_consensus(
                    results,
                    title
                )
            )

            if best:

                row.update({
                    "consensus_status":
                        "Result found",

                    "consensus_title":
                        best.get(
                            "title",
                            ""
                        ),

                    "consensus_url":
                        best.get(
                            "url",
                            ""
                        ),

                    "consensus_match_score":
                        f"{score:.3f}",
                })

                print(
                    f"  Consensus match: "
                    f"{best.get('title', '')}"
                )

                print(
                    f"  Match score: "
                    f"{score:.3f}"
                )

                # Official Consensus search responses normally
                # provide the paper URL rather than a PDF URL.
                # If an API/integration exposes a public PDF URL,
                # try it here.
                for pdf_url in (
                    consensus_public_pdf_urls(
                        best
                    )
                ):

                    ok, message, final_url = (
                        download_pdf(
                            pdf_url,
                            output_path
                        )
                    )

                    if ok:

                        row.update({
                            "local_pdf":
                                str(output_path),

                            "final_source":
                                "Consensus",

                            "download_status":
                                "Downloaded - Consensus",
                        })

                        print(
                            "  ✓ PDF downloaded "
                            "via Consensus."
                        )

                        break

                    row["error"] = (
                        f"{row['error']} | "
                        f"Consensus PDF: "
                        f"{message}"
                    ).strip(
                        " |"
                    )

            else:

                row[
                    "consensus_status"
                ] = "No matching result"

        else:

            if CONSENSUS_API_KEY:

                row[
                    "consensus_status"
                ] = "No API result"

            else:

                row[
                    "consensus_status"
                ] = (
                    "Manual link only - "
                    "API key not configured"
                )

                print(
                    "  Consensus API key not configured."
                )

                print(
                    "  Search URL saved in report."
                )

        if row["local_pdf"]:

            rows.append(
                row
            )

            write_csv(
                rows,
                REPORT_DIR /
                "master_paper_database_progress.csv"
            )

            continue

        # ====================================================
        # STAGE 3: GOOGLE SCHOLAR
        # ====================================================

        print()
        print(
            "  [3/4] Google Scholar..."
        )

        scholar_result = (
            scholar_find_and_download(
                doi,
                title,
                output_path
            )
        )

        row.update({
            "scholar_status":
                scholar_result.get(
                    "status",
                    ""
                ),

            "scholar_result_title":
                scholar_result.get(
                    "result_title",
                    ""
                ),

            "scholar_result_url":
                scholar_result.get(
                    "result_url",
                    ""
                ),

            "scholar_pdf_url":
                scholar_result.get(
                    "pdf_url",
                    ""
                ),
        })

        if (
            scholar_result.get(
                "found"
            )
        ):

            if (
                scholar_result.get(
                    "status",
                    ""
                ).startswith(
                    "Downloaded"
                )
            ):

                row.update({
                    "local_pdf":
                        str(output_path),

                    "final_source":
                        "Google Scholar",

                    "download_status":
                        "Downloaded - Google Scholar",
                })

            else:

                row[
                    "download_status"
                ] = (
                    "Scholar result - "
                    "no downloadable PDF"
                )

        else:

            row[
                "download_status"
            ] = (
                "Scholar result not found"
            )

        time.sleep(
            SCHOLAR_DELAY
        )

        if row["local_pdf"]:

            rows.append(
                row
            )

            write_csv(
                rows,
                REPORT_DIR /
                "master_paper_database_progress.csv"
            )

            continue

        # ====================================================
        # STAGE 4: OPENALEX LEGAL OA
        # ====================================================

        if ENABLE_OPENALEX:

            print()
            print(
                "  [4/4] Extended legal OA "
                "search - OpenAlex..."
            )

            openalex, oa_error = (
                openalex_search(
                    doi
                )
            )

            if openalex:

                row[
                    "openalex_status"
                ] = "Record found"

                for pdf_url in (
                    openalex_pdf_candidates(
                        openalex
                    )
                ):

                    ok, message, final_url = (
                        download_pdf(
                            pdf_url,
                            output_path
                        )
                    )

                    if ok:

                        row.update({
                            "openalex_pdf_url":
                                pdf_url,

                            "local_pdf":
                                str(output_path),

                            "final_source":
                                "OpenAlex",

                            "download_status":
                                "Downloaded - OpenAlex",
                        })

                        print(
                            "  ✓ PDF downloaded "
                            "from OpenAlex."
                        )

                        break

                    row["error"] = (
                        f"{row['error']} | "
                        f"OpenAlex PDF: "
                        f"{message}"
                    ).strip(
                        " |"
                    )

            else:

                row[
                    "openalex_status"
                ] = "No record"

                if oa_error:

                    row["error"] = (
                        f"{row['error']} | "
                        f"OpenAlex: "
                        f"{oa_error}"
                    ).strip(
                        " |"
                    )

        if not row["local_pdf"]:

            row[
                "download_status"
            ] = (
                "Manual review required"
            )

            print(
                "  ⚠ No legal/public PDF found."
            )

        # ====================================================
        # SAVE AFTER EVERY PAPER
        # ====================================================

        rows.append(
            row
        )

        write_csv(
            rows,
            REPORT_DIR /
            "master_paper_database_progress.csv"
        )

    # ========================================================
    # FINAL REPORTS
    # ========================================================

    master_csv = (
        REPORT_DIR /
        "master_paper_database.csv"
    )

    master_xlsx = (
        REPORT_DIR /
        "master_paper_database.xlsx"
    )

    write_csv(
        rows,
        master_csv
    )

    excel_created = write_excel(
        rows,
        master_xlsx
    )

    total = len(rows)

    existing = sum(
        r.get(
            "existing_pdf"
        ) == "Yes"
        for r in rows
    )

    downloaded = sum(
        r.get(
            "download_status",
            ""
        ).startswith(
            "Downloaded"
        )
        for r in rows
    )

    unpaywall = sum(
        r.get(
            "final_source"
        ) == "Unpaywall"
        for r in rows
    )

    consensus = sum(
        r.get(
            "final_source"
        ) == "Consensus"
        for r in rows
    )

    scholar = sum(
        r.get(
            "final_source"
        ) == "Google Scholar"
        for r in rows
    )

    openalex = sum(
        r.get(
            "final_source"
        ) == "OpenAlex"
        for r in rows
    )

    unresolved = sum(
        not r.get(
            "local_pdf",
            ""
        )
        for r in rows
    )

    summary = f"""
DOI RESEARCH ARTICLE DOWNLOADER V2.2
====================================

MASTER / ONLY INPUT:
{SCOPUS_EXCEL}

DOI SOURCE:
Scopus Excel only

SCOPUS SHEET:
{SCOPUS_SHEET}

PAPER NUMBER:
Scopus S.No (permanent PDF/file ID)

RETRIEVAL ORDER:
Existing -> Unpaywall -> Consensus -> Google Scholar
-> OpenAlex -> Manual Review

TOTAL EXCEL DOI RECORDS:
{total}

EXISTING PDFs PRESERVED:
{existing}

NEW PDFs DOWNLOADED:
{downloaded}

Unpaywall:
{unpaywall}

Consensus:
{consensus}

Google Scholar:
{scholar}

OpenAlex:
{openalex}

UNRESOLVED / MANUAL:
{unresolved}

DOIs NOT FOUND IN SCOPUS MASTER:
{len(unmatched_scopus)}

EXCEL REPORT CREATED:
{excel_created}

PDF DIRECTORY:
{PDF_DIR.resolve()}

REPORT DIRECTORY:
{REPORT_DIR.resolve()}

IMPORTANT:
- Existing PDFs are not deleted.
- Existing PDFs are not deliberately re-downloaded.
- New PDFs use the Scopus S.No as the permanent prefix.
- Google Scholar [PDF] links in div.gs_ggs are explicitly checked.
- Scholar PDFs are downloaded before the browser is closed.
- No CAPTCHA, paywall, login, or access control is bypassed.
"""

    (
        REPORT_DIR /
        "SUMMARY.txt"
    ).write_text(
        summary.strip()
        + "\n",
        encoding="utf-8"
    )

    if unmatched_scopus:

        (
            REPORT_DIR /
            "DOIs_not_found_in_Scopus.txt"
        ).write_text(
            "\n".join(
                unmatched_scopus
            )
            + "\n",
            encoding="utf-8"
        )

    print()
    print("=" * 72)
    print(
        "V2.2 COMPLETED"
    )
    print("=" * 72)

    print(
        f"Total records          : {total}"
    )

    print(
        f"Existing PDFs preserved: {existing}"
    )

    print(
        f"New PDFs downloaded    : {downloaded}"
    )

    print(
        f"  Unpaywall            : {unpaywall}"
    )

    print(
        f"  Consensus            : {consensus}"
    )

    print(
        f"  Google Scholar       : {scholar}"
    )

    print(
        f"  OpenAlex             : {openalex}"
    )

    print(
        f"Manual/unresolved      : {unresolved}"
    )

    print(
        f"DOIs absent from Scopus: "
        f"{len(unmatched_scopus)}"
    )

    print()
    print(
        "Reports:"
    )

    print(
        REPORT_DIR.resolve()
    )


if __name__ == "__main__":
    main()
